import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
import io
import re
import logging
import sys
from decimal import Decimal, ROUND_HALF_UP

# Configure logging
def setup_logging():
    """Set up logging configuration for the Streamlit app"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(filename)s:%(lineno)d - %(funcName)s() - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout),  # Console output
            logging.FileHandler('myntra_pricing.log')  # File output
        ]
    )
    return logging.getLogger(__name__)

# Initialize logger
logger = setup_logging()

# Import functions from your existing myntra_pricing.py
def parse_excel_if(formula: str):
    formula = formula.strip().lstrip("=").replace('\n', "").replace(" ", "")
    pattern = re.compile(r"IF\([^\<]*\<(\d+),(\d+),?")
    matches = pattern.findall(formula)
    
    results = []
    for threshold, value in matches:
        results.append((int(threshold), int(value)))
    
    else_value_match = re.search(r",(\d+)\)*$", formula)
    if else_value_match:
        results.append((float('inf'), int(else_value_match.group(1))))
    
    return results

def dynamic_if(rules, value):
    for threshold, result in rules:
        if value < threshold:
            return result
    return rules[-1][1]

def calc_customer_shipping_charges(formula, value):
    rules = parse_excel_if(formula)
    result = dynamic_if(rules, value)
    return result

def calc_commission_charges(formula, value):
    rules = parse_excel_if(formula)
    result = dynamic_if(rules, value)
    return result

def calc_fixed_fee(formula, value):
    rules = parse_excel_if(formula)
    result = dynamic_if(rules, value)
    return result

def round_selling_price_myntra(price, endings=None):
    """
    Round a selling price to end in one of the specified endings (default: [49, 99]).
    If the next higher candidate is within 10, round up to it.
    Otherwise round down to the nearest candidate.
    """
    if endings is None:
        endings = [49, 99]
    if price <= 0:
        return price

    hundred = int(price / 100) * 100
    candidates = []
    for h in [hundred - 100, hundred, hundred + 100]:
        for end in endings:
            val = h + end
            if val > 0:
                candidates.append(val)

    lower_candidates = [c for c in candidates if c <= price]
    higher_candidates = [c for c in candidates if c >= price]

    lower = max(lower_candidates) if lower_candidates else None
    higher = min(higher_candidates) if higher_candidates else None

    if lower is None:
        return higher
    if higher is None:
        return lower
    if (higher - price) <= (price - lower):
        return higher
    return lower

def profit_percent_from_discount_myntra(discount, df, show_details=False, price_endings=None):
    """Calculate profit for Myntra portal"""
    try:
        # Safely extract and convert values, handling text-formatted numbers
        mrp = safe_convert_to_numeric(df['MRP'], 'MRP', 0)
        cp = safe_convert_to_numeric(df['CP'], 'CP', 0)
        gst = safe_convert_to_numeric(df['GST'], 'GST', 0)
        customer_shipping_charges_formula = df['SHIPPING']
        commission_formula = df['COMMISSION %']
        fixed_fee_formula = df['FIXED FEE']
        mbb_td = safe_convert_to_numeric(df.get('REBATE TD', 0), 'REBATE TD', 0)
        mbb_rebate_pct = safe_convert_to_numeric(df.get('REBATE VALUE', 0), 'REBATE VALUE', 0)
        mbb_rsp = safe_convert_to_numeric(df.get('REBATE SP', 0), 'REBATE SP', 0)

        # Validate essential values
        if pd.isna(mrp) or mrp <= 0:
            logger.warning(f"Invalid MRP value: {df['MRP']} (converted to {mrp})")
            return 0, 0
        if pd.isna(cp) or cp <= 0:
            logger.warning(f"Invalid CP value: {df['CP']} (converted to {cp})")
            return 0, 0

        selling_price = round_selling_price_myntra(mrp - (mrp * discount / 100), endings=price_endings)

        customer_shipping_charges = calc_customer_shipping_charges(customer_shipping_charges_formula, selling_price)
        selling_price_after_log = selling_price - customer_shipping_charges
        gst_amount = selling_price * gst / 100
        commission_percent = calc_commission_charges(commission_formula, selling_price_after_log)
        commission_amount = selling_price_after_log * commission_percent / 100
        fixed_fee = calc_fixed_fee(fixed_fee_formula, selling_price_after_log)
        commission_gst = commission_amount * 0.18
        fixed_fee_gst = fixed_fee * 0.18
        gross_settlement = selling_price_after_log - (commission_amount + commission_gst) - (fixed_fee + fixed_fee_gst)
        return_fee = selling_price_after_log * 0.02
        marketting_packing_cost = selling_price_after_log * 0.1
        rebate_qualifies = (mbb_rsp > 0 and selling_price < mbb_rsp) or (mbb_td > 0 and discount >= mbb_td)
        mbb_rebate = (mbb_rebate_pct / 100 * selling_price_after_log) if rebate_qualifies else 0
        profit = gross_settlement - cp - gst_amount - return_fee - marketting_packing_cost + mbb_rebate
        profit_percent = profit / selling_price * 100

        if show_details:
            return {
                'profit': profit,
                'profit_percent': profit_percent,
                'details': {
                    'mrp': mrp,
                    'discount': discount,
                    'selling_price': selling_price,
                    'customer_shipping_charges': customer_shipping_charges,
                    'selling_price_after_log': selling_price_after_log,
                    'gst': gst,
                    'gst_amount': gst_amount,
                    'commission_percent': commission_percent,
                    'commission_amount': commission_amount,
                    'commission_gst': commission_gst,
                    'fixed_fee': fixed_fee,
                    'fixed_fee_gst': fixed_fee_gst,
                    'gross_settlement': gross_settlement,
                    'return_fee': return_fee,
                    'marketting_packing_cost': marketting_packing_cost,
                    'mbb_td': mbb_td,
                    'mbb_rebate_pct': mbb_rebate_pct,
                    'mbb_rebate': mbb_rebate,
                    'cp': cp,
                    'profit': profit,
                    'profit_percent': profit_percent
                }
            }
        
        return profit, profit_percent

    except Exception as e:
        logger.error(f"Error in Myntra profit calculation: {str(e)} | Discount: {discount} | MRP: {df.get('MRP', 'N/A')} | CP: {df.get('CP', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0, 0

def profit_percent_from_selling_price_myntra(selling_price, df, show_details=False):
    """Calculate Myntra profit for an explicit selling price."""
    try:
        mrp = safe_convert_to_numeric(df['MRP'], 'MRP', 0)
        cp = safe_convert_to_numeric(df['CP'], 'CP', 0)
        gst = safe_convert_to_numeric(df['GST'], 'GST', 0)
        customer_shipping_charges_formula = df['SHIPPING']
        commission_formula = df['COMMISSION %']
        fixed_fee_formula = df['FIXED FEE']
        mbb_td = safe_convert_to_numeric(df.get('REBATE TD', 0), 'REBATE TD', 0)
        mbb_rebate_pct = safe_convert_to_numeric(df.get('REBATE VALUE', 0), 'REBATE VALUE', 0)
        mbb_rsp = safe_convert_to_numeric(df.get('REBATE SP', 0), 'REBATE SP', 0)

        if pd.isna(mrp) or mrp <= 0:
            logger.warning(f"Invalid MRP value: {df['MRP']} (converted to {mrp})")
            return 0, 0
        if pd.isna(cp) or cp <= 0:
            logger.warning(f"Invalid CP value: {df['CP']} (converted to {cp})")
            return 0, 0
        if pd.isna(selling_price) or selling_price <= 0:
            return 0, 0

        discount = ((mrp - selling_price) / mrp) * 100 if mrp > 0 else 0
        customer_shipping_charges = calc_customer_shipping_charges(customer_shipping_charges_formula, selling_price)
        selling_price_after_log = selling_price - customer_shipping_charges
        gst_amount = selling_price * gst / 100
        commission_percent = calc_commission_charges(commission_formula, selling_price_after_log)
        commission_amount = selling_price_after_log * commission_percent / 100
        fixed_fee = calc_fixed_fee(fixed_fee_formula, selling_price_after_log)
        commission_gst = commission_amount * 0.18
        fixed_fee_gst = fixed_fee * 0.18
        gross_settlement = selling_price_after_log - (commission_amount + commission_gst) - (fixed_fee + fixed_fee_gst)
        return_fee = selling_price_after_log * 0.02
        marketting_packing_cost = selling_price_after_log * 0.1
        rebate_qualifies = (mbb_rsp > 0 and selling_price < mbb_rsp) or (mbb_td > 0 and discount >= mbb_td)
        mbb_rebate = (mbb_rebate_pct / 100 * selling_price_after_log) if rebate_qualifies else 0
        profit = gross_settlement - cp - gst_amount - return_fee - marketting_packing_cost + mbb_rebate
        profit_percent = profit / selling_price * 100

        if show_details:
            return {
                'profit': profit,
                'profit_percent': profit_percent,
                'details': {
                    'mrp': mrp,
                    'discount': discount,
                    'selling_price': selling_price,
                    'customer_shipping_charges': customer_shipping_charges,
                    'selling_price_after_log': selling_price_after_log,
                    'gst': gst,
                    'gst_amount': gst_amount,
                    'commission_percent': commission_percent,
                    'commission_amount': commission_amount,
                    'commission_gst': commission_gst,
                    'fixed_fee': fixed_fee,
                    'fixed_fee_gst': fixed_fee_gst,
                    'gross_settlement': gross_settlement,
                    'return_fee': return_fee,
                    'marketting_packing_cost': marketting_packing_cost,
                    'mbb_td': mbb_td,
                    'mbb_rebate_pct': mbb_rebate_pct,
                    'mbb_rebate': mbb_rebate,
                    'cp': cp,
                    'profit': profit,
                    'profit_percent': profit_percent
                }
            }

        return profit, profit_percent

    except Exception as e:
        logger.error(f"Error in Myntra selling price calculation: {str(e)} | Selling Price: {selling_price} | MRP: {df.get('MRP', 'N/A')} | CP: {df.get('CP', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0, 0

def find_optimal_mrp_myntra(discount, target_profit_percent, min_absolute_profit, df, show_details=False, price_endings=None):
    """
    Find optimal MRP for Myntra portal given constant discount and target profit percentage.
    Iterates on MRP values to find the MRP that achieves the desired profit.
    
    Args:
        discount: Fixed discount percentage to apply
        target_profit_percent: Target profit percentage to achieve
        min_absolute_profit: Minimum absolute profit amount required
        df: Product data row
        show_details: Whether to return detailed calculation breakdown
    
    Returns:
        tuple: (optimal_mrp, profit, profit_percent) or (None, 0, 0) if no solution found
    """
    try:
        # Safely extract and convert values, handling text-formatted numbers
        cp = safe_convert_to_numeric(df['CP'], 'CP', 0)
        gst = safe_convert_to_numeric(df['GST'], 'GST', 0)
        customer_shipping_charges_formula = df['SHIPPING']
        commission_formula = df['COMMISSION %']
        fixed_fee_formula = df['FIXED FEE']
        mbb_td = safe_convert_to_numeric(df.get('REBATE TD', 0), 'REBATE TD', 0)
        mbb_rebate_pct = safe_convert_to_numeric(df.get('REBATE VALUE', 0), 'REBATE VALUE', 0)
        mbb_rsp = safe_convert_to_numeric(df.get('REBATE SP', 0), 'REBATE SP', 0)

        # Validate essential values
        if pd.isna(cp) or cp <= 0:
            logger.warning(f"Invalid CP value: {df['CP']} (converted to {cp})")
            return None, 0, 0

        # Start with a reasonable MRP based on CP, rounded to nearest 100
        start_mrp = round(cp * 2.5 / 100) * 100
        best_mrp = None
        best_profit = 0
        best_profit_percent = 0
        
        # Search range: from 1.5x CP to 50x CP, in steps of 100
        min_mrp = max(100, round(cp * 1.5 / 100) * 100)  # At least ₹100
        max_mrp = round(cp * 50 / 100) * 100  # Up to 50x CP
        
        # Linear search approach with steps of 100
        # This is necessary because IF conditions in formulas can cause non-linear behavior
        tolerance = 0  # 0.5% tolerance for profit percentage (more lenient for linear search)
        
        # Iterate through MRP values in steps of 100
        for test_mrp in range(int(min_mrp), int(max_mrp) + 100, 100):
            # Calculate profit for this MRP
            selling_price = round_selling_price_myntra(test_mrp - (test_mrp * discount / 100), endings=price_endings)

            customer_shipping_charges = calc_customer_shipping_charges(customer_shipping_charges_formula, selling_price)
            selling_price_after_log = selling_price - customer_shipping_charges
            gst_amount = selling_price * gst / 100
            commission_percent = calc_commission_charges(commission_formula, selling_price_after_log)
            commission_amount = selling_price_after_log * commission_percent / 100
            fixed_fee = calc_fixed_fee(fixed_fee_formula, selling_price_after_log)
            commission_gst = commission_amount * 0.18
            fixed_fee_gst = fixed_fee * 0.18
            gross_settlement = selling_price_after_log - (commission_amount + commission_gst) - (fixed_fee + fixed_fee_gst)
            return_fee = selling_price_after_log * 0.02
            marketting_packing_cost = selling_price_after_log * 0.1
            rebate_qualifies = (mbb_rsp > 0 and selling_price < mbb_rsp) or (mbb_td > 0 and discount >= mbb_td)
            mbb_rebate = (mbb_rebate_pct / 100 * selling_price_after_log) if rebate_qualifies else 0
            profit = gross_settlement - cp - gst_amount - return_fee - marketting_packing_cost + mbb_rebate
            profit_percent = profit / selling_price * 100 if selling_price > 0 else 0

            # Check if this meets our criteria
            if (profit_percent >= target_profit_percent - tolerance and 
                profit >= min_absolute_profit and 
                profit_percent <= target_profit_percent + tolerance):
                # Found exact match within tolerance
                best_mrp = test_mrp
                best_profit = profit
                best_profit_percent = profit_percent
                break
                
            # Keep track of best solution so far (meets minimum criteria)
            if (profit_percent >= target_profit_percent and 
                profit >= min_absolute_profit and 
                (best_mrp is None or abs(profit_percent - target_profit_percent) < abs(best_profit_percent - target_profit_percent))):
                best_mrp = test_mrp
                best_profit = profit
                best_profit_percent = profit_percent
        
        if best_mrp is None:
            logger.warning(f"No optimal MRP found for discount {discount}%, target profit {target_profit_percent}%")
            return None, 0, 0
        
        if show_details:
            # Recalculate with best MRP for detailed breakdown
            selling_price = round_selling_price_myntra(best_mrp - (best_mrp * discount / 100), endings=price_endings)
            customer_shipping_charges = calc_customer_shipping_charges(customer_shipping_charges_formula, selling_price)
            selling_price_after_log = selling_price - customer_shipping_charges
            gst_amount = selling_price * gst / 100
            commission_percent = calc_commission_charges(commission_formula, selling_price_after_log)
            commission_amount = selling_price_after_log * commission_percent / 100
            fixed_fee = calc_fixed_fee(fixed_fee_formula, selling_price_after_log)
            commission_gst = commission_amount * 0.18
            fixed_fee_gst = fixed_fee * 0.18
            gross_settlement = selling_price_after_log - (commission_amount + commission_gst) - (fixed_fee + fixed_fee_gst)
            return_fee = selling_price_after_log * 0.02
            marketting_packing_cost = selling_price_after_log * 0.1
            rebate_qualifies = (mbb_rsp > 0 and selling_price < mbb_rsp) or (mbb_td > 0 and discount >= mbb_td)
            mbb_rebate = (mbb_rebate_pct / 100 * selling_price_after_log) if rebate_qualifies else 0
            profit = gross_settlement - cp - gst_amount - return_fee - marketting_packing_cost + mbb_rebate
            profit_percent = profit / selling_price * 100

            return {
                'optimal_mrp': best_mrp,
                'profit': profit,
                'profit_percent': profit_percent,
                'details': {
                    'mrp': best_mrp,
                    'discount': discount,
                    'selling_price': selling_price,
                    'customer_shipping_charges': customer_shipping_charges,
                    'selling_price_after_log': selling_price_after_log,
                    'gst': gst,
                    'gst_amount': gst_amount,
                    'commission_percent': commission_percent,
                    'commission_amount': commission_amount,
                    'commission_gst': commission_gst,
                    'fixed_fee': fixed_fee,
                    'fixed_fee_gst': fixed_fee_gst,
                    'gross_settlement': gross_settlement,
                    'return_fee': return_fee,
                    'marketting_packing_cost': marketting_packing_cost,
                    'mbb_td': mbb_td,
                    'mbb_rebate_pct': mbb_rebate_pct,
                    'mbb_rebate': mbb_rebate,
                    'cp': cp,
                    'profit': profit,
                    'profit_percent': profit_percent
                }
            }
        
        return best_mrp, best_profit, best_profit_percent

    except Exception as e:
        logger.error(f"Error in Myntra MRP calculation: {str(e)} | Discount: {discount} | Target Profit: {target_profit_percent}% | CP: {df.get('CP', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return None, 0, 0

def profit_percent_from_discount_ajio(discount, df, all_cost_percent=42, show_details=False):
    """Calculate profit for Ajio portal"""
    try:
        # Safely extract and convert values, handling text-formatted numbers
        mrp = safe_convert_to_numeric(df['Listing MRP'], 'Listing MRP', 0)
        if pd.isna(mrp) or mrp <= 0:
            logger.warning(f"Invalid Listing MRP value: {df['Listing MRP']} (converted to {mrp})")
            return 0, 0

        selling_price = mrp - (mrp * discount / 100)
        return profit_percent_from_selling_price_ajio(selling_price, df, all_cost_percent, show_details)

    except Exception as e:
        logger.error(f"Error in Ajio profit calculation: {str(e)} | Discount: {discount} | MRP: {df.get('Listing MRP', 'N/A')} | CP: {df.get('CP', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0, 0

def profit_percent_from_selling_price_ajio(selling_price, df, all_cost_percent=42, show_details=False):
    """Calculate Ajio profit for an explicit selling price."""
    try:
        mrp = safe_convert_to_numeric(df['Listing MRP'], 'Listing MRP', 0)
        cp = safe_convert_to_numeric(df['CP'], 'CP', 0)
        gst = safe_convert_to_numeric(df['GST'], 'GST', 0)

        if pd.isna(mrp) or mrp <= 0:
            logger.warning(f"Invalid Listing MRP value: {df['Listing MRP']} (converted to {mrp})")
            return 0, 0
        if pd.isna(cp) or cp <= 0:
            logger.warning(f"Invalid CP value: {df['CP']} (converted to {cp})")
            return 0, 0
        if pd.isna(selling_price) or selling_price <= 0:
            return 0, 0

        discount = ((mrp - selling_price) / mrp) * 100 if mrp > 0 else 0
        gst_value = selling_price * gst / 100
        comission = selling_price * 0.35
        comission_gst = comission * 0.18
        total_comission = comission + comission_gst
        marketting = selling_price * 0.1

        total_cost = marketting + total_comission + gst_value + cp
        
        profit = selling_price - total_cost
        profit_percent = profit / selling_price * 100
        
        if show_details:
            return {
                'profit': profit,
                'profit_percent': profit_percent,
                'details': {
                    'mrp': mrp,
                    'discount': discount,
                    'selling_price': selling_price,
                    'all_cost_percent': all_cost_percent,
                    'gst_value': gst_value,
                    'comission': comission,
                    'comission_gst': comission_gst,
                    'total_comission': total_comission,
                    'marketting': marketting,
                    'total_cost': total_cost,
                    'cp': cp,
                    'profit': profit,
                    'profit_percent': profit_percent
                }
            }
        
        return profit, profit_percent

    except Exception as e:
        logger.error(f"Error in Ajio selling price calculation: {str(e)} | Selling Price: {selling_price} | MRP: {df.get('Listing MRP', 'N/A')} | CP: {df.get('CP', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0, 0

def round_price_to_end_with_nine(price):
    """
    Round a price to end with 9.
    Examples: 327.45 -> 329, 458.12 -> 459, 198.67 -> 199
    
    Args:
        price (float): The original price
        
    Returns:
        float: Price rounded to end with 9
    """
    if price <= 0:
        return price
    
    # Match Excel ROUND(price, -1) behavior, then subtract 1 to end with 9.
    rounded_to_ten = float((Decimal(str(price)) / Decimal('10')).quantize(Decimal('1'), rounding=ROUND_HALF_UP) * Decimal('10'))
    price_ending_nine = rounded_to_ten - 1
    
    # Handle edge case where rounding would make price too low
    if price_ending_nine <= 0:
        price_ending_nine = 9
    
    return price_ending_nine

def calculate_tatacliq_profit_from_selling_price(selling_price, df, show_details=False):
    """Calculate profit for TataCliq portal from a given selling price"""
    try:
        # Safely extract and convert values, handling text-formatted numbers
        mrp = safe_convert_to_numeric(df['MRP'], 'MRP', 0)
        cp = safe_convert_to_numeric(df['CP'], 'CP', 0)
        gst_rate = safe_convert_to_numeric(df['GST RATE'], 'GST RATE', 0)

        # Validate essential values
        if pd.isna(mrp) or mrp <= 0:
            logger.warning(f"Invalid MRP value: {df['MRP']} (converted to {mrp})")
            return 0, 0
        if pd.isna(cp) or cp <= 0:
            logger.warning(f"Invalid CP value: {df['CP']} (converted to {cp})")
            return 0, 0

        # Calculate discount from selling price
        discount = ((mrp - selling_price) / mrp) * 100 if mrp > 0 else 0

        gst_value = gst_rate * selling_price / 100
        processing_fee = safe_convert_to_numeric(df.get('Processing fee', 49), 'Processing fee', 49)
        processing_fee_gst = processing_fee * 0.18

        if selling_price < 500:
            shipping_charge = 0
            commission = 150
        else:
            shipping_charge = 118
            commission = 0.25 * selling_price

        igst = 0.18 * commission 
        total_fees = commission + igst 
        
        marketting_fees = 0.05 * selling_price
        total_cost = gst_value + shipping_charge + total_fees + cp + marketting_fees + processing_fee + processing_fee_gst

        profit = selling_price - total_cost
        profit_percent = profit / selling_price * 100 if selling_price > 0 else 0

        if show_details:
            return {
                'profit': profit,
                'profit_percent': profit_percent,
                'details': {
                    'mrp': mrp,
                    'discount': discount,
                    'selling_price': selling_price,
                    'gst_rate': gst_rate,
                    'gst_value': gst_value,
                    'commission': commission,
                    'igst': igst,
                    'total_fees': total_fees,
                    'shipping_charge': shipping_charge,
                    'marketting_fees': marketting_fees,
                    'processing_fee': processing_fee,
                    'processing_fee_gst': processing_fee_gst,
                    'cp': cp,
                    'total_cost': total_cost,
                    'profit': profit,
                    'profit_percent': profit_percent
                }
            }
        
        return profit, profit_percent

    except Exception as e:
        logger.error(f"Error in TataCliq profit calculation from selling price: {str(e)} | Selling Price: {selling_price} | MRP: {df.get('MRP', 'N/A')} | CP: {df.get('CP', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0, 0

def profit_percent_from_discount_tatacliq_with_nine_ending(discount, df, show_details=False):
    """Calculate profit for TataCliq portal with selling price ending in 9 (legacy function for compatibility)"""
    try:
        # Safely extract and convert values, handling text-formatted numbers
        mrp = safe_convert_to_numeric(df['MRP'], 'MRP', 0)
        cp = safe_convert_to_numeric(df['CP'], 'CP', 0)

        # Validate essential values
        if pd.isna(mrp) or mrp <= 0:
            logger.warning(f"Invalid MRP value: {df['MRP']} (converted to {mrp})")
            return 0, 0
        if pd.isna(cp) or cp <= 0:
            logger.warning(f"Invalid CP value: {df['CP']} (converted to {cp})")
            return 0, 0

        # Calculate initial selling price
        initial_selling_price = mrp - (mrp * discount / 100)
        
        # Round selling price to end with 9
        selling_price = round_price_to_end_with_nine(initial_selling_price)

        # Use the new function to calculate profit from selling price
        return calculate_tatacliq_profit_from_selling_price(selling_price, df, show_details)

    except Exception as e:
        logger.error(f"Error in TataCliq profit calculation (nine-ending): {str(e)} | Discount: {discount} | MRP: {df.get('MRP', 'N/A')} | CP: {df.get('CP', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0, 0

def profit_percent_from_discount_tatacliq(discount, df, show_details=False):
    """Calculate profit for TataCliq portal"""
    try:
        # Safely extract and convert values, handling text-formatted numbers
        mrp = safe_convert_to_numeric(df['MRP'], 'MRP', 0)
        cp = safe_convert_to_numeric(df['CP'], 'CP', 0)
        gst_rate = safe_convert_to_numeric(df['GST RATE'], 'GST RATE', 0)

        # Validate essential values
        if pd.isna(mrp) or mrp <= 0:
            logger.warning(f"Invalid MRP value: {df['MRP']} (converted to {mrp})")
            return 0, 0
        if pd.isna(cp) or cp <= 0:
            logger.warning(f"Invalid CP value: {df['CP']} (converted to {cp})")
            return 0, 0

        selling_price = round_price_to_end_with_nine(mrp - (mrp * discount / 100))

        gst_value = gst_rate * selling_price / 100
        processing_fee = safe_convert_to_numeric(df.get('Processing fee', 49), 'Processing fee', 49)
        processing_fee_gst = processing_fee * 0.18

        if selling_price < 500:
            shipping_charge = 0
            commission = 150
        else:
            shipping_charge = 118
            commission = 0.25 * selling_price

        igst = 0.18 * commission 
        total_fees = commission + igst 
        
        marketting_fees = 0.05 * selling_price
        total_cost = gst_value + shipping_charge + total_fees + cp + marketting_fees + processing_fee + processing_fee_gst

        profit = selling_price - total_cost
        profit_percent = profit / selling_price * 100

        if show_details:
            return {
                'profit': profit,
                'profit_percent': profit_percent,
                'details': {
                    'mrp': mrp,
                    'discount': discount,
                    'selling_price': selling_price,
                    'gst_rate': gst_rate,
                    'gst_value': gst_value,
                    'commission': commission,
                    'igst': igst,
                    'total_fees': total_fees,
                    'shipping_charge': shipping_charge,
                    'marketting_fees': marketting_fees,
                    'processing_fee': processing_fee,
                    'processing_fee_gst': processing_fee_gst,
                    'cp': cp,
                    'total_cost': total_cost,
                    'profit': profit,
                    'profit_percent': profit_percent
                }
            }
        
        return profit, profit_percent

    except Exception as e:
        logger.error(f"Error in TataCliq profit calculation: {str(e)} | Discount: {discount} | MRP: {df.get('MRP', 'N/A')} | CP: {df.get('CP', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0, 0

def profit_percent_from_discount_nykaa(discount, df, show_details=False):
    """Calculate profit for Nykaa portal"""
    try:
        # Safely extract and convert values, handling text-formatted numbers
        mrp = safe_convert_to_numeric(df['MRP'], 'MRP', 0)
        if pd.isna(mrp) or mrp <= 0:
            logger.warning(f"Invalid MRP value: {df['MRP']} (converted to {mrp})")
            return 0, 0

        selling_price = mrp - (mrp * discount / 100)
        return profit_percent_from_selling_price_nykaa(selling_price, df, show_details)

    except Exception as e:
        logger.error(f"Error in Nykaa profit calculation: {str(e)} | Discount: {discount} | MRP: {df.get('MRP', 'N/A')} | CP: {df.get('cp', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0, 0

def profit_percent_from_selling_price_nykaa(selling_price, df, show_details=False):
    """Calculate Nykaa profit for an explicit selling price."""
    try:
        mrp = safe_convert_to_numeric(df['MRP'], 'MRP', 0)
        cp = safe_convert_to_numeric(df['cp'], 'cp', 0)
        gst_rate = safe_convert_to_numeric(df['gst'], 'gst', 0)
        shipping = safe_convert_to_numeric(df['shipping'], 'shipping', 0)

        if pd.isna(mrp) or mrp <= 0:
            logger.warning(f"Invalid MRP value: {df['MRP']} (converted to {mrp})")
            return 0, 0
        if pd.isna(cp) or cp <= 0:
            logger.warning(f"Invalid CP value: {df['cp']} (converted to {cp})")
            return 0, 0
        if pd.isna(selling_price) or selling_price <= 0:
            return 0, 0

        discount = ((mrp - selling_price) / mrp) * 100 if mrp > 0 else 0
        gst_value = gst_rate * selling_price / 100
        
        commission = 0.28 * selling_price
        commission_tax = 0.18 * commission
        total_commission = commission + commission_tax
        
        marketing_fees = 0.1 * selling_price
        payment_fee = selling_price * 0.008
        payment_fee_gst = payment_fee * 0.18
        total_cost = cp + gst_value + shipping + marketing_fees + total_commission + payment_fee + payment_fee_gst

        profit = selling_price - total_cost
        profit_percent = profit / selling_price * 100

        if show_details:
            return {
                'profit': profit,
                'profit_percent': profit_percent,
                'details': {
                    'mrp': mrp,
                    'discount': discount,
                    'selling_price': selling_price,
                    'gst_rate': gst_rate,
                    'gst_value': gst_value,
                    'commission': commission,
                    'commission_tax': commission_tax,
                    'total_commission': total_commission,
                    'shipping': shipping,
                    'marketing_fees': marketing_fees,
                    'payment_fee': payment_fee,
                    'payment_fee_gst': payment_fee_gst,
                    'cp': cp,
                    'total_cost': total_cost,
                    'profit': profit,
                    'profit_percent': profit_percent
                }
            }
        
        return profit, profit_percent

    except Exception as e:
        logger.error(f"Error in Nykaa selling price calculation: {str(e)} | Selling Price: {selling_price} | MRP: {df.get('MRP', 'N/A')} | CP: {df.get('cp', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0, 0

def profit_percent_from_discount_pepperfry(discount, df, all_cost_percent=42, show_details=False):
    """Calculate profit for Pepperfry portal"""
    try:
        # Safely extract and convert values, handling text-formatted numbers
        mrp = safe_convert_to_numeric(df['mrp'], 'mrp', 0)
        if pd.isna(mrp) or mrp <= 0:
            logger.warning(f"Invalid Listing MRP value: {df['mrp']} (converted to {mrp})")
            return 0, 0

        selling_price = mrp - (mrp * discount / 100)
        return profit_percent_from_selling_price_pepperfry(selling_price, df, all_cost_percent, show_details)

    except Exception as e:
        logger.error(f"Error in Pepperfry profit calculation: {str(e)} | Discount: {discount} | MRP: {df.get('mrp', 'N/A')} | CP: {df.get('cp', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0, 0

def profit_percent_from_selling_price_pepperfry(selling_price, df, all_cost_percent=42, show_details=False):
    """Calculate Pepperfry profit for an explicit selling price."""
    try:
        mrp = safe_convert_to_numeric(df['mrp'], 'mrp', 0)
        cp = safe_convert_to_numeric(df['cp'], 'cp', 0)
        gst = safe_convert_to_numeric(df['gst'], 'gst', 0)

        if pd.isna(mrp) or mrp <= 0:
            logger.warning(f"Invalid Listing MRP value: {df['mrp']} (converted to {mrp})")
            return 0, 0
        if pd.isna(cp) or cp <= 0:
            logger.warning(f"Invalid CP value: {df['cp']} (converted to {cp})")
            return 0, 0
        if pd.isna(selling_price) or selling_price <= 0:
            return 0, 0

        discount = ((mrp - selling_price) / mrp) * 100 if mrp > 0 else 0
        gst_value = selling_price * gst / 100
        comission = selling_price * 0.35
        comission_gst = comission * 0
        total_comission = comission + comission_gst
        marketting = selling_price * 0.1

        total_cost = marketting + total_comission + gst_value + cp
        
        profit = selling_price - total_cost
        profit_percent = profit / selling_price * 100
        
        if show_details:
            return {
                'profit': profit,
                'profit_percent': profit_percent,
                'details': {
                    'mrp': mrp,
                    'discount': discount,
                    'selling_price': selling_price,
                    'all_cost_percent': all_cost_percent,
                    'gst_value': gst_value,
                    'comission': comission,
                    'comission_gst': comission_gst,
                    'total_comission': total_comission,
                    'marketting': marketting,
                    'total_cost': total_cost,
                    'cp': cp,
                    'profit': profit,
                    'profit_percent': profit_percent
                }
            }
        
        return profit, profit_percent

    except Exception as e:
        logger.error(f"Error in Pepperfry selling price calculation: {str(e)} | Selling Price: {selling_price} | MRP: {df.get('mrp', 'N/A')} | CP: {df.get('cp', 'N/A')}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return 0, 0

def get_profit_calculation_function(portal, use_nine_ending=False):
    """Return the appropriate profit calculation function based on portal and pricing strategy"""
    if portal == 'TataCliq' and use_nine_ending:
        return profit_percent_from_discount_tatacliq_with_nine_ending
    
    portal_functions = {
        'Myntra': profit_percent_from_discount_myntra,
        'Ajio': profit_percent_from_discount_ajio,
        'TataCliq': profit_percent_from_discount_tatacliq,
        'Nykaa': profit_percent_from_discount_nykaa,
        'Pepperfry': profit_percent_from_discount_pepperfry
    }
    return portal_functions.get(portal, profit_percent_from_discount_myntra)

def get_portal_mrp(row, portal):
    """Return the MRP/listing price column used to generate candidate selling prices."""
    mrp_columns = {
        'Myntra': 'MRP',
        'Ajio': 'Listing MRP',
        'TataCliq': 'MRP',
        'Nykaa': 'MRP',
        'Pepperfry': 'mrp',
    }
    column = mrp_columns.get(portal, 'MRP')
    return safe_convert_to_numeric(row.get(column, 0), column, 0)

def generate_nine_ending_prices(mrp):
    """Generate selling prices ending in 9, from low to high, up to MRP."""
    if pd.isna(mrp) or mrp <= 0:
        return []
    return list(range(9, int(mrp) + 1, 10))

def calculate_profit_from_selling_price(portal, selling_price, row, show_details=False, **kwargs):
    """Calculate portal profit from an explicit selling price."""
    if portal == 'Myntra':
        return profit_percent_from_selling_price_myntra(selling_price, row, show_details=show_details)
    if portal == 'Ajio':
        return profit_percent_from_selling_price_ajio(selling_price, row, kwargs.get('all_cost_percent', 42), show_details=show_details)
    if portal == 'TataCliq':
        return calculate_tatacliq_profit_from_selling_price(selling_price, row, show_details=show_details)
    if portal == 'Nykaa':
        return profit_percent_from_selling_price_nykaa(selling_price, row, show_details=show_details)
    if portal == 'Pepperfry':
        return profit_percent_from_selling_price_pepperfry(selling_price, row, kwargs.get('all_cost_percent', 42), show_details=show_details)
    return profit_percent_from_selling_price_myntra(selling_price, row, show_details=show_details)

def display_detailed_calculations(df, portal, result_df, **kwargs):
    """Display detailed calculations for the first 2 rows using their best 9-ending selling price."""
    
    st.info("This section shows the detailed calculation breakdown for the first 2 products using their optimal 9-ending selling prices.")
    
    for row_idx in range(min(2, len(df))):
        row = df.iloc[row_idx]
        row_index = df.index[row_idx]
        
        # Get the best selling price and its equivalent discount from the result dataframe
        if row_idx < len(result_df):
            best_discount = result_df.iloc[row_idx]['Best Discount']
            best_selling_price = result_df.iloc[row_idx].get('Best Selling Price')
            
            if pd.isna(best_selling_price) or best_selling_price is None:
                st.markdown(f"### Row {row_idx + 1} - {row_index}")
                st.warning("No suitable selling price found for this product")
                st.markdown("---")
                continue
        else:
            st.markdown(f"### Row {row_idx + 1} - {row_index}")
            st.warning("No result data available for this row")
            st.markdown("---")
            continue
        
        st.markdown(f"### Row {row_idx + 1} - {row_index}")
        st.info(f"**Best Selling Price: ₹{best_selling_price:.0f}** | **Equivalent Discount: {best_discount:.2f}%**")
        
        try:
            result = calculate_profit_from_selling_price(portal, best_selling_price, row, show_details=True, **kwargs)
            
            if result and 'details' in result:
                details = result['details']
                
                # Create columns for better layout
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**Input Variables:**")
                    for key, value in details.items():
                        if key in ['mrp', 'cp', 'gst', 'discount', 'all_cost_percent']:
                            if key == 'discount':
                                st.write(f"• **{key.replace('_', ' ').title()}:** {value}%")
                            else:
                                st.write(f"• **{key.replace('_', ' ').title()}:** {value}")
                
                with col2:
                    st.markdown("**Calculated Values:**")
                    for key, value in details.items():
                        if key not in ['mrp', 'cp', 'gst', 'discount', 'all_cost_percent']:
                            if isinstance(value, (int, float)):
                                st.write(f"• **{key.replace('_', ' ').title()}:** ₹{value:.2f}")
                            else:
                                st.write(f"• **{key.replace('_', ' ').title()}:** {value}")
                
                # Summary
                st.markdown("---")
                if 'gross_settlement' in details:
                    col1, col2, col3, col4, col5 = st.columns(5)
                    with col1:
                        st.metric("Selling Price", f"₹{details['selling_price']:.2f}")
                    with col2:
                        st.metric("Gross Settlement", f"₹{details['gross_settlement']:.2f}")
                    with col3:
                        st.metric("Profit", f"₹{result['profit']:.2f}")
                    with col4:
                        st.metric("Profit %", f"{result['profit_percent']:.2f}%")
                    with col5:
                        st.metric("Eq. Discount", f"{details['discount']:.2f}%")
                else:
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Selling Price", f"₹{details['selling_price']:.2f}")
                    with col2:
                        st.metric("Profit", f"₹{result['profit']:.2f}")
                    with col3:
                        st.metric("Profit %", f"{result['profit_percent']:.2f}%")
                    with col4:
                        st.metric("Eq. Discount", f"{details['discount']:.2f}%")

        except Exception as e:
            logger.error(f"Error in detailed calculation for row {row_idx}: {str(e)} | Selling Price: {best_selling_price} | Portal: {portal}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            st.error(f"Error calculating for ₹{best_selling_price:.0f} selling price: {str(e)}")
        
        st.markdown("---")

def display_detailed_calculations_for_article(df, portal, result_df, **kwargs):
    """Display detailed calculations for a user-specified article number."""
    calculation_mode = kwargs.get('calculation_mode', 'discount')

    article_no = st.text_input(
        "Enter Article Number",
        placeholder="e.g. 12345678",
        help="Type the Article Number exactly as it appears in your data"
    )

    if not article_no:
        st.info("Enter an Article Number above to see its detailed calculation breakdown.")
        return

    # Try to match as-is, then as integer if numeric
    index_vals = df.index.astype(str)
    match = article_no.strip() in index_vals

    if not match:
        st.error(f"Article Number **{article_no}** not found in the processed data.")
        return

    # Locate the row (use string-matched index)
    row = df.loc[df.index.astype(str) == article_no.strip()].iloc[0]
    row_label = row.name

    st.markdown(f"### {row_label}")

    if calculation_mode == 'selling_price':
        minimum_sp = result_df.loc[result_df.index.astype(str) == article_no.strip(), 'Minimum Selling Price']
        minimum_sp = minimum_sp.iloc[0] if len(minimum_sp) else None

        if pd.isna(minimum_sp) or minimum_sp is None:
            st.warning("No suitable selling price found for this product.")
            return

        st.info(f"**Minimum Selling Price: ₹{int(minimum_sp)}**")

        try:
            result = profit_percent_from_selling_price_myntra(int(minimum_sp), row, show_details=True)
            if result and 'details' in result:
                display_calculation_details(result, f"₹{int(minimum_sp)}", "Minimum SP")
        except Exception as e:
            logger.error(f"Error in article lookup selling price calculation: {str(e)} | Article: {article_no} | Selling Price: {minimum_sp}")
            st.error(f"Error calculating details: {str(e)}")
        return

    best_discount = result_df.loc[result_df.index.astype(str) == article_no.strip(), 'Best Discount']
    best_discount = best_discount.iloc[0] if len(best_discount) else None
    best_selling_price = result_df.loc[result_df.index.astype(str) == article_no.strip(), 'Best Selling Price']
    best_selling_price = best_selling_price.iloc[0] if len(best_selling_price) else None

    if pd.isna(best_selling_price) or best_selling_price is None:
        st.warning("No suitable selling price found for this product.")
        return

    st.info(f"**Best Selling Price: ₹{best_selling_price:.0f}** | **Equivalent Discount: {best_discount:.2f}%**")

    try:
        result = calculate_profit_from_selling_price(portal, best_selling_price, row, show_details=True, **kwargs)

        if result and 'details' in result:
            details = result['details']

            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**Input Variables:**")
                for key, value in details.items():
                    if key in ['mrp', 'cp', 'gst', 'discount', 'all_cost_percent']:
                        if key == 'discount':
                            st.write(f"• **{key.replace('_', ' ').title()}:** {value}%")
                        else:
                            st.write(f"• **{key.replace('_', ' ').title()}:** {value}")

            with col2:
                st.markdown("**Calculated Values:**")
                for key, value in details.items():
                    if key not in ['mrp', 'cp', 'gst', 'discount', 'all_cost_percent']:
                        if isinstance(value, (int, float)):
                            st.write(f"• **{key.replace('_', ' ').title()}:** ₹{value:.2f}")
                        else:
                            st.write(f"• **{key.replace('_', ' ').title()}:** {value}")

            st.markdown("---")
            if 'gross_settlement' in details:
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric("Selling Price", f"₹{details['selling_price']:.2f}")
                with col2:
                    st.metric("Gross Settlement", f"₹{details['gross_settlement']:.2f}")
                with col3:
                    st.metric("Profit", f"₹{result['profit']:.2f}")
                with col4:
                    st.metric("Profit %", f"{result['profit_percent']:.2f}%")
                with col5:
                    st.metric("Eq. Discount", f"{details['discount']:.2f}%")
            else:
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Selling Price", f"₹{details['selling_price']:.2f}")
                with col2:
                    st.metric("Profit", f"₹{result['profit']:.2f}")
                with col3:
                    st.metric("Profit %", f"{result['profit_percent']:.2f}%")
                with col4:
                    st.metric("Eq. Discount", f"{details['discount']:.2f}%")

    except Exception as e:
        logger.error(f"Error in article lookup calculation: {str(e)} | Article: {article_no} | Selling Price: {best_selling_price}")
        st.error(f"Error calculating details: {str(e)}")

def display_calculation_details(result, primary_value, primary_label):
    """Display a compact two-column calculation breakdown."""
    details = result['details']

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Input Variables:**")
        for key, value in details.items():
            if key in ['mrp', 'cp', 'gst', 'discount', 'all_cost_percent']:
                if key == 'discount':
                    st.write(f"• **{key.replace('_', ' ').title()}:** {value:.2f}%")
                else:
                    st.write(f"• **{key.replace('_', ' ').title()}:** {value}")

    with col2:
        st.markdown("**Calculated Values:**")
        for key, value in details.items():
            if key not in ['mrp', 'cp', 'gst', 'discount', 'all_cost_percent']:
                if isinstance(value, (int, float)):
                    st.write(f"• **{key.replace('_', ' ').title()}:** ₹{value:.2f}")
                else:
                    st.write(f"• **{key.replace('_', ' ').title()}:** {value}")

    st.markdown("---")
    if 'gross_settlement' in details:
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric(primary_label, primary_value)
        with col2:
            st.metric("Gross Settlement", f"₹{details['gross_settlement']:.2f}")
        with col3:
            st.metric("Profit", f"₹{result['profit']:.2f}")
        with col4:
            st.metric("Profit %", f"{result['profit_percent']:.2f}%")
        with col5:
            st.metric("Selling Price", f"₹{details['selling_price']:.2f}")
    else:
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric(primary_label, primary_value)
        with col2:
            st.metric("Profit", f"₹{result['profit']:.2f}")
        with col3:
            st.metric("Profit %", f"{result['profit_percent']:.2f}%")
        with col4:
            st.metric("Selling Price", f"₹{details['selling_price']:.2f}")

def display_detailed_selling_price_calculations(df, portal, result_df):
    """Display detailed calculations for the first 2 rows using their minimum selling prices."""
    st.info("This section shows the detailed calculation breakdown for the first 2 products using their minimum qualifying selling prices.")

    for row_idx in range(min(2, len(df))):
        row = df.iloc[row_idx]
        row_index = df.index[row_idx]

        if row_idx < len(result_df):
            minimum_sp = result_df.iloc[row_idx]['Minimum Selling Price']
            status = result_df.iloc[row_idx]['Status']

            if pd.isna(minimum_sp) or minimum_sp is None or status != 'Found':
                st.markdown(f"### Row {row_idx + 1} - {row_index}")
                st.warning(f"No minimum selling price found for this product (Status: {status})")
                st.markdown("---")
                continue
        else:
            st.markdown(f"### Row {row_idx + 1} - {row_index}")
            st.warning("No result data available for this row")
            st.markdown("---")
            continue

        st.markdown(f"### Row {row_idx + 1} - {row_index}")
        st.info(f"**Minimum Selling Price: ₹{int(minimum_sp)}**")

        try:
            if portal == 'Myntra':
                result = profit_percent_from_selling_price_myntra(int(minimum_sp), row, show_details=True)
                if result and 'details' in result:
                    display_calculation_details(result, f"₹{int(minimum_sp)}", "Minimum SP")
        except Exception as e:
            logger.error(f"Error in detailed selling price calculation for row {row_idx}: {str(e)} | Selling Price: ₹{minimum_sp} | Portal: {portal}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            st.error(f"Error calculating for selling price ₹{minimum_sp}: {str(e)}")

        st.markdown("---")


def display_detailed_mrp_calculations(df, portal, result_df, **kwargs):
    """Display detailed calculations for the first 2 rows using their optimal MRP"""
    st.info("This section shows the detailed calculation breakdown for the first 2 products using their optimal MRP values.")
    
    for row_idx in range(min(15, len(df))):
        row = df.iloc[row_idx]
        row_index = df.index[row_idx]
        
        # Get the optimal MRP from the result dataframe
        if row_idx < len(result_df):
            optimal_mrp = result_df.iloc[row_idx]['Optimal MRP']
            status = result_df.iloc[row_idx]['Status']
            
            if pd.isna(optimal_mrp) or optimal_mrp is None or status != 'Found':
                st.markdown(f"### Row {row_idx + 1} - {row_index}")
                st.warning(f"No optimal MRP found for this product (Status: {status})")
                st.markdown("---")
                continue
        else:
            st.markdown(f"### Row {row_idx + 1} - {row_index}")
            st.warning("No result data available for this row")
            st.markdown("---")
            continue
        
        st.markdown(f"### Row {row_idx + 1} - {row_index}")
        st.info(f"**Optimal MRP: ₹{optimal_mrp:.0f}**")
        
        try:
            # Get parameters from kwargs
            discount = kwargs.get('discount', 20)
            target_profit_percent = kwargs.get('target_profit_percent', 15)
            min_absolute_profit = kwargs.get('min_absolute_profit', 100)
            
            # Calculate detailed breakdown
            result = find_optimal_mrp_myntra(
                discount, target_profit_percent, min_absolute_profit, row, show_details=True
            )
            
            if result and 'details' in result:
                details = result['details']
                
                # Create columns for better layout
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**Input Variables:**")
                    st.write(f"• **Cost Price (CP):** ₹{details['cp']:.2f}")
                    st.write(f"• **GST Rate:** {details['gst']:.1f}%")
                    st.write(f"• **Fixed Discount:** {details['discount']:.1f}%")
                    st.write(f"• **Target Profit:** {target_profit_percent:.1f}%")
                    st.write(f"• **Min Absolute Profit:** ₹{min_absolute_profit:.0f}")
                
                with col2:
                    st.markdown("**Calculated Values:**")
                    st.write(f"• **Optimal MRP:** ₹{details['mrp']:.0f}")
                    st.write(f"• **Selling Price:** ₹{details['selling_price']:.2f}")
                    st.write(f"• **Customer Shipping:** ₹{details['customer_shipping_charges']:.2f}")
                    st.write(f"• **Selling Price After Logistics:** ₹{details['selling_price_after_log']:.2f}")
                    st.write(f"• **GST Amount:** ₹{details['gst_amount']:.2f}")
                    st.write(f"• **Commission %:** {details['commission_percent']:.2f}%")
                    st.write(f"• **Commission Amount:** ₹{details['commission_amount']:.2f}")
                    st.write(f"• **Commission GST (18%):** ₹{details['commission_gst']:.2f}")
                    st.write(f"• **Fixed Fee:** ₹{details['fixed_fee']:.2f}")
                    st.write(f"• **Fixed Fee GST (18%):** ₹{details['fixed_fee_gst']:.2f}")
                    st.write(f"• **Gross Settlement:** ₹{details['gross_settlement']:.2f}")
                    st.write(f"• **Return Fee:** ₹{details['return_fee']:.2f}")
                    st.write(f"• **Marketing & Packing:** ₹{details['marketting_packing_cost']:.2f}")
                
                # Summary
                st.markdown("---")
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric("Optimal MRP", f"₹{details['mrp']:.0f}")
                with col2:
                    st.metric("Gross Settlement", f"₹{details['gross_settlement']:.2f}")
                with col3:
                    st.metric("Profit", f"₹{result['profit']:.2f}")
                with col4:
                    st.metric("Profit %", f"{result['profit_percent']:.2f}%")
                with col5:
                    st.metric("Selling Price", f"₹{details['selling_price']:.2f}")
                
                # Additional info
                st.markdown("---")
                st.markdown("**Calculation Summary:**")
                st.write(f"• **MRP Search Range:** ₹{max(100, round(details['cp'] * 1.5 / 100) * 100):.0f} to ₹{round(details['cp'] * 50 / 100) * 100:.0f} (in steps of ₹100)")
                st.write(f"• **Fixed Discount Applied:** {discount:.1f}%")
                st.write(f"• **Target Achieved:** {'✅' if result['profit_percent'] >= target_profit_percent and result['profit'] >= min_absolute_profit else '❌'}")
            
        except Exception as e:
            logger.error(f"Error in detailed MRP calculation for row {row_idx}: {str(e)} | MRP: ₹{optimal_mrp:.0f} | Portal: {portal}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            st.error(f"Error calculating for MRP ₹{optimal_mrp:.0f}: {str(e)}")
        
        st.markdown("---")

def build_profit_table(df, target_profit_percent, min_absolute_profit, portal, **kwargs):
    profit_data = []
    profit_data_abs = []
    
    logger.info(f"Starting profit table build for {portal} with {len(df)} rows")
    logger.info(f"Target profit: {target_profit_percent}%, Min absolute profit: {min_absolute_profit}")
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_rows = len(df)
    
    for idx, (index, row) in enumerate(df.iterrows()):
        # Update progress every 10 rows to reduce logging overhead
        if idx % 10 == 0 or idx == total_rows - 1:
            status_text.text(f'Processing row {idx + 1} of {total_rows} for {portal}...')
            progress_bar.progress((idx + 1) / total_rows)
        
        row_profit = {}
        row_profit_abs = {}
        best_discount = None
        best_profit = 0
        best_profit_percent = 0
        best_selling_price = None
        best_gross_settlement = None
        prev_profit = None
        weird_jumps = []

        mbb_td_row = safe_convert_to_numeric(row.get('REBATE TD', 0), 'REBATE TD', 0) if portal == 'Myntra' else 0
        mbb_rsp_row = safe_convert_to_numeric(row.get('REBATE SP', 0), 'REBATE SP', 0) if portal == 'Myntra' else 0

        mrp_row = get_portal_mrp(row, portal)
        selling_prices = generate_nine_ending_prices(mrp_row)

        for selling_price in selling_prices:
            try:
                result = calculate_profit_from_selling_price(portal, selling_price, row, show_details=True, **kwargs)
                if isinstance(result, dict):
                    profit = result['profit']
                    profit_pct = result['profit_percent']
                    details = result['details']
                    discount = details.get('discount', ((mrp_row - selling_price) / mrp_row * 100) if mrp_row > 0 else 0)
                    gross_settlement = details.get('gross_settlement')
                else:
                    profit, profit_pct = result
                    discount = ((mrp_row - selling_price) / mrp_row * 100) if mrp_row > 0 else 0
                    gross_settlement = None

                price_key = f'₹{int(selling_price)}'
                row_profit[price_key] = round(profit_pct, 2)
                row_profit_abs[price_key] = round(profit, 2)

                if profit_pct >= target_profit_percent and profit > min_absolute_profit:
                    if best_selling_price is None or selling_price < best_selling_price:
                        best_discount = discount
                        best_profit = profit
                        best_profit_percent = profit_pct
                        best_selling_price = selling_price
                        best_gross_settlement = gross_settlement

                if prev_profit is not None and profit_pct > 15 and profit_pct > prev_profit + 0.01 and discount != mbb_td_row:
                    weird_jumps.append(selling_price)
                prev_profit = profit_pct

            except Exception as e:
                logger.warning(f"Error calculating profit for selling price ₹{selling_price} in row {idx}: {str(e)} | Row data: {dict(row)}")
                row_profit[f'₹{int(selling_price)}'] = None
                row_profit_abs[f'₹{int(selling_price)}'] = None

        # Profit at the REBATE TD threshold discount
        if portal == 'Myntra' and mbb_td_row > 0:
            try:
                rebate_selling_price = round_price_to_end_with_nine(mrp_row - (mrp_row * mbb_td_row / 100))
                mbb_profit_abs, mbb_profit_pct = calculate_profit_from_selling_price(portal, rebate_selling_price, row, **kwargs)
                row_profit['REBATE TD Discount'] = int(mbb_td_row)
                row_profit['Profit at REBATE TD'] = round(mbb_profit_abs, 2)
                row_profit['Profit % at REBATE TD'] = round(mbb_profit_pct, 2)
            except Exception as e:
                logger.warning(f"Error calculating profit at MBB TD for row {idx}: {str(e)}")
                row_profit['REBATE TD Discount'] = int(mbb_td_row)
                row_profit['Profit at REBATE TD'] = None
                row_profit['Profit % at REBATE TD'] = None

        row_profit['Best Discount'] = best_discount
        row_profit['Best Selling Price'] = round(best_selling_price, 2) if best_selling_price is not None else None
        row_profit['Gross Settlement'] = round(best_gross_settlement, 2) if best_gross_settlement is not None else None
        row_profit['Best Profit (₹)'] = best_profit
        row_profit['Best Profit %'] = round(best_profit_percent, 2) if best_selling_price is not None else None
        row_profit['Weird Profit Jump'] = ','.join(str(x) for x in weird_jumps)
        profit_data.append(row_profit)

        row_profit_abs['Best Discount'] = best_discount
        row_profit_abs['Best Selling Price'] = round(best_selling_price, 2) if best_selling_price is not None else None
        row_profit_abs['Gross Settlement'] = round(best_gross_settlement, 2) if best_gross_settlement is not None else None
        row_profit_abs['Best Profit (₹)'] = best_profit
        row_profit_abs['Best Profit %'] = round(best_profit_percent, 2) if best_selling_price is not None else None
        profit_data_abs.append(row_profit_abs)

    progress_bar.empty()
    status_text.empty()
    
    logger.info(f"Completed processing {len(profit_data)} rows for {portal}")
    
    # Log summary statistics
    products_with_target = sum(1 for row in profit_data if row.get('Best Selling Price') is not None)
    logger.info(f"Summary: {products_with_target}/{len(profit_data)} products met target profit criteria")
    
    def _reorder(df_in):
        cols = df_in.columns.tolist()
        for c in ['Best Profit %', 'Best Profit (₹)', 'Gross Settlement', 'Best Discount', 'Best Selling Price']:
            if c in cols:
                cols.insert(0, cols.pop(cols.index(c)))
        for mbb_col in ['Profit % at REBATE TD', 'Profit at REBATE TD', 'REBATE TD Discount']:
            if mbb_col in cols:
                cols.insert(0, cols.pop(cols.index(mbb_col)))
        return df_in[cols]

    combined_df = _reorder(pd.DataFrame(profit_data, index=df.index))
    combined_abs_df = _reorder(pd.DataFrame(profit_data_abs, index=df.index))

    return combined_df, combined_abs_df

def build_profit_table_nine_ending(df, target_profit_percent, min_absolute_profit, portal, **kwargs):
    """Optimized profit table build for TataCliq nine-ending pricing strategy"""
    profit_data = []
    
    logger.info(f"Starting optimized nine-ending profit table build for {portal} with {len(df)} rows")
    logger.info(f"Target profit: {target_profit_percent}%, Min absolute profit: {min_absolute_profit}")
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_rows = len(df)
    
    for idx, (index, row) in enumerate(df.iterrows()):
        # Update progress every 10 rows to reduce logging overhead
        if idx % 10 == 0 or idx == total_rows - 1:
            status_text.text(f'Processing row {idx + 1} of {total_rows} for {portal} (nine-ending)...')
            progress_bar.progress((idx + 1) / total_rows)
        
        try:
            mrp = safe_convert_to_numeric(row['MRP'], 'MRP', 0)
            
            if pd.isna(mrp) or mrp <= 0:
                # Skip invalid rows
                profit_data.append({
                    'Best Discount': None,
                    'Best Profit (₹)': 0,
                    'Weird Profit Jump': '',
                    'Best Selling Price': None
                })
                continue
            
            row_profit = {}
            best_discount = None
            best_profit = 0
            best_selling_price = None
            prev_profit = None
            weird_jumps = []
            
            # Iterate through selling prices ending in 9: 9, 19, 29, 39, ..., up to MRP
            for selling_price in range(9, int(mrp) + 1, 10):
                try:
                    # Calculate profit from selling price
                    profit, profit_pct = calculate_tatacliq_profit_from_selling_price(selling_price, row)
                    
                    # Calculate discount for display
                    discount = ((mrp - selling_price) / mrp) * 100 if mrp > 0 else 0
                    
                    # Store profit percentage for this selling price (using discount as key for consistency)
                    discount_key = f'{int(discount)}%'
                    row_profit[discount_key] = round(profit_pct, 2)
                    
                    # Check if this meets our criteria
                    if profit_pct >= target_profit_percent and profit > min_absolute_profit:
                        best_discount = int(discount)
                        best_profit = profit
                        best_selling_price = selling_price

                    # Check for weird jumps
                    if prev_profit is not None and profit_pct > 15 and profit_pct > prev_profit + 0.01:
                        weird_jumps.append(int(discount))
                    prev_profit = profit_pct

                except Exception as e:
                    logger.warning(f"Error calculating profit for selling price {selling_price} in row {idx}: {str(e)}")
                    discount = ((mrp - selling_price) / mrp) * 100 if mrp > 0 else 0
                    row_profit[f'{int(discount)}%'] = None

            row_profit['Best Discount'] = best_discount
            row_profit['Best Profit (₹)'] = best_profit
            row_profit['Best Selling Price'] = best_selling_price
            row_profit['Weird Profit Jump'] = ','.join(str(x) for x in weird_jumps)
            profit_data.append(row_profit)
            
        except Exception as e:
            logger.warning(f"Error processing row {idx}: {str(e)} | Row data: {dict(row)}")
            profit_data.append({
                'Best Discount': None,
                'Best Profit (₹)': 0,
                'Best Selling Price': None,
                'Weird Profit Jump': ''
            })

    progress_bar.empty()
    status_text.empty()
    
    logger.info(f"Completed optimized processing {len(profit_data)} rows for {portal}")
    
    # Log summary statistics
    products_with_target = sum(1 for row in profit_data if row.get('Best Discount') is not None)
    logger.info(f"Summary: {products_with_target}/{len(profit_data)} products met target profit criteria")
    
    combined_df = pd.DataFrame(profit_data, index=df.index)
    cols = combined_df.columns.tolist()

    priority_cols = ['Best Profit (₹)', 'Best Discount', 'Best Selling Price']
    for col in priority_cols:
        if col in cols:
            cols.insert(0, cols.pop(cols.index(col)))

    combined_df = combined_df[cols]

    return combined_df, None

def build_mrp_table(df, discount, target_profit_percent, min_absolute_profit, portal, **kwargs):
    """Build table for MRP calculation mode (constant discount, find optimal MRP)"""
    mrp_data = []
    
    logger.info(f"Starting MRP table build for {portal} with {len(df)} rows")
    logger.info(f"Fixed discount: {discount}%, Target profit: {target_profit_percent}%, Min absolute profit: {min_absolute_profit}")
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_rows = len(df)
    
    for idx, (index, row) in enumerate(df.iterrows()):
        # Update progress every 10 rows to reduce logging overhead
        if idx % 10 == 0 or idx == total_rows - 1:
            status_text.text(f'Processing row {idx + 1} of {total_rows} for {portal} MRP calculation...')
            progress_bar.progress((idx + 1) / total_rows)
        
        row_mrp = {}
        
        try:
            if portal == 'Myntra':
                optimal_mrp, profit, profit_pct = find_optimal_mrp_myntra(
                    discount, target_profit_percent, min_absolute_profit, row,
                    price_endings=kwargs.get('price_endings', None)
                )
                
                row_mrp['Optimal MRP'] = round(optimal_mrp, 2) if optimal_mrp else None
                row_mrp['Profit'] = round(profit, 2)
                row_mrp['Profit %'] = round(profit_pct, 2)
                row_mrp['Selling Price'] = round_selling_price_myntra(optimal_mrp - (optimal_mrp * discount / 100), endings=kwargs.get('price_endings', None)) if optimal_mrp else None
                row_mrp['Gross Settlement'] = None
                if row_mrp['Selling Price'] is not None:
                    detail_result = profit_percent_from_selling_price_myntra(row_mrp['Selling Price'], row, show_details=True)
                    if isinstance(detail_result, dict):
                        row_mrp['Gross Settlement'] = round(detail_result['details'].get('gross_settlement', 0), 2)
                row_mrp['Status'] = 'Found' if optimal_mrp else 'No Solution'
                
            else:
                # For other portals, we can extend this later
                row_mrp['Optimal MRP'] = None
                row_mrp['Profit'] = 0
                row_mrp['Profit %'] = 0
                row_mrp['Selling Price'] = None
                row_mrp['Gross Settlement'] = None
                row_mrp['Status'] = 'Not Supported'
                
        except Exception as e:
            logger.warning(f"Error calculating MRP for row {idx}: {str(e)} | Row data: {dict(row)}")
            row_mrp['Optimal MRP'] = None
            row_mrp['Profit'] = 0
            row_mrp['Profit %'] = 0
            row_mrp['Selling Price'] = None
            row_mrp['Gross Settlement'] = None
            row_mrp['Status'] = 'Error'
        
        mrp_data.append(row_mrp)

    progress_bar.empty()
    status_text.empty()
    
    logger.info(f"Completed processing {len(mrp_data)} rows for {portal} MRP calculation")
    
    # Log summary statistics
    products_with_solution = sum(1 for row in mrp_data if row.get('Status') == 'Found')
    logger.info(f"Summary: {products_with_solution}/{len(mrp_data)} products found optimal MRP solutions")
    
    combined_df = pd.DataFrame(mrp_data, index=df.index)
    
    return combined_df

def build_selling_price_table(df, target_profit_percent, min_absolute_profit, portal, **kwargs):
    """Build table by scanning selected selling price endings and picking the minimum viable SP."""
    sp_data = []
    selected_endings = kwargs.get('price_endings') or [9, 19, 29, 39, 49, 59, 69, 79, 89, 99]
    selected_endings = sorted({int(ending) for ending in selected_endings})

    logger.info(f"Starting selling price table build for {portal} with {len(df)} rows")
    logger.info(f"Target profit: {target_profit_percent}%, Min absolute profit: {min_absolute_profit}")
    logger.info(f"Selected selling price endings: {selected_endings}")

    progress_bar = st.progress(0)
    status_text = st.empty()

    total_rows = len(df)

    for idx, (index, row) in enumerate(df.iterrows()):
        if idx % 10 == 0 or idx == total_rows - 1:
            status_text.text(f'Processing row {idx + 1} of {total_rows} for {portal} selling price calculation...')
            progress_bar.progress((idx + 1) / total_rows)

        row_sp = {}

        try:
            if portal != 'Myntra':
                row_sp['Minimum Selling Price'] = None
                row_sp['Equivalent Discount'] = None
                row_sp['Gross Settlement'] = None
                row_sp['Profit'] = 0
                row_sp['Profit %'] = 0
                row_sp['Status'] = 'Not Supported'
                sp_data.append(row_sp)
                continue

            mrp = safe_convert_to_numeric(row['MRP'], 'MRP', 0)
            if pd.isna(mrp) or mrp <= 0:
                row_sp['Minimum Selling Price'] = None
                row_sp['Equivalent Discount'] = None
                row_sp['Gross Settlement'] = None
                row_sp['Profit'] = 0
                row_sp['Profit %'] = 0
                row_sp['Status'] = 'Invalid MRP'
                sp_data.append(row_sp)
                continue

            minimum_sp = None
            best_profit = 0
            best_profit_pct = 0
            best_gross_settlement = None
            equivalent_discount = None

            candidate_prices = sorted(
                price
                for base in range(0, int(mrp) + 100, 100)
                for ending in selected_endings
                if 0 < (price := base + ending) <= mrp
            )

            for selling_price in candidate_prices:
                profit, profit_pct = profit_percent_from_selling_price_myntra(selling_price, row)

                if profit_pct >= target_profit_percent and profit >= min_absolute_profit:
                    minimum_sp = selling_price
                    best_profit = profit
                    best_profit_pct = profit_pct
                    equivalent_discount = ((mrp - selling_price) / mrp) * 100
                    detail_result = profit_percent_from_selling_price_myntra(selling_price, row, show_details=True)
                    if isinstance(detail_result, dict):
                        best_gross_settlement = detail_result['details'].get('gross_settlement')
                    break

            row_sp['Minimum Selling Price'] = minimum_sp
            row_sp['Equivalent Discount'] = round(equivalent_discount, 2) if equivalent_discount is not None else None
            row_sp['Gross Settlement'] = round(best_gross_settlement, 2) if best_gross_settlement is not None else None
            row_sp['Profit'] = round(best_profit, 2)
            row_sp['Profit %'] = round(best_profit_pct, 2)
            row_sp['Status'] = 'Found' if minimum_sp is not None else 'No Solution'

        except Exception as e:
            logger.warning(f"Error calculating selling price for row {idx}: {str(e)} | Row data: {dict(row)}")
            row_sp['Minimum Selling Price'] = None
            row_sp['Equivalent Discount'] = None
            row_sp['Gross Settlement'] = None
            row_sp['Profit'] = 0
            row_sp['Profit %'] = 0
            row_sp['Status'] = 'Error'

        sp_data.append(row_sp)

    progress_bar.empty()
    status_text.empty()

    logger.info(f"Completed processing {len(sp_data)} rows for {portal} selling price calculation")
    products_with_solution = sum(1 for row in sp_data if row.get('Status') == 'Found')
    logger.info(f"Summary: {products_with_solution}/{len(sp_data)} products found minimum selling price solutions")

    return pd.DataFrame(sp_data, index=df.index)

def safe_convert_to_numeric(value, column_name="", default_value=0):
    """
    Safely convert a value to numeric, handling text-formatted numbers from Excel.
    
    Args:
        value: The value to convert
        column_name: Name of the column for error reporting
        default_value: Default value to return if conversion fails
    
    Returns:
        Numeric value or default_value if conversion fails
    """
    # Handle Series objects by taking the first value
    if isinstance(value, pd.Series):
        if len(value) == 0:
            return default_value
        value = value.iloc[0]
    
    # Handle None and NaN values
    if value is None:
        return default_value
    
    try:
        if pd.isna(value):
            return default_value
    except (ValueError, TypeError):
        # If pd.isna fails, treat as non-null and continue
        pass
    
    # If already numeric, return as is
    if isinstance(value, (int, float)):
        return value
    
    # Convert string to numeric
    if isinstance(value, str):
        # Remove common text formatting that might interfere
        cleaned_value = value.strip().replace(',', '').replace('₹', '').replace('$', '').replace('%', '')
        
        # Handle empty strings
        if not cleaned_value:
            return default_value
        
        try:
            # Try to convert to float first, then int if it's a whole number
            numeric_value = float(cleaned_value)
            if numeric_value.is_integer():
                return int(numeric_value)
            return numeric_value
        except (ValueError, TypeError) as e:
            logger.warning(f"Could not convert '{value}' to numeric in column '{column_name}': {str(e)}. Using default value {default_value}")
            return default_value
    
    # For other types, try direct conversion
    try:
        # Handle numpy types
        if hasattr(value, 'item'):
            value = value.item()
        return float(value)
    except (ValueError, TypeError, AttributeError) as e:
        logger.warning(f"Could not convert {type(value).__name__} '{value}' to numeric in column '{column_name}': {str(e)}. Using default value {default_value}")
        return default_value

def validate_and_convert_dataframe(df, portal, required_columns):
    """
    Validate and convert DataFrame columns to appropriate data types.
    
    Args:
        df: DataFrame to process
        portal: Portal name for error reporting
        required_columns: List of required columns for the portal
    
    Returns:
        Processed DataFrame with converted data types
    """
    logger.info(f"Validating and converting data types for {portal}")
    logger.info(f"DataFrame columns: {list(df.columns)} | Index: {df.index.names}")
    logger.info(f"Required columns: {required_columns}")
    
    # Create a copy to avoid modifying original
    df_processed = df.copy()
    
    # Define numeric columns for each portal (exclude identifier columns like SKU Code, ARTICLE NO, EAN)
    numeric_columns = {
        'Myntra': ['MRP', 'CP', 'GST'],
        'Ajio': ['CP', 'Listing MRP', 'GST'],
        'TataCliq': ['CP', 'MRP', 'GST RATE', 'Processing fee'],
        'Nykaa': ['MRP', 'cp', 'gst', 'shipping'],
        'Pepperfry': ['cp', 'mrp', 'gst']
    }
    
    # Define identifier columns that should remain as text
    identifier_columns = ['SKU Code', 'ARTICLE NO', 'EAN', 'SKU', 'Product Code', 'Item Code', 'van']
    
    # Get numeric columns for current portal
    portal_numeric_cols = numeric_columns.get(portal, [])
    
    # Convert numeric columns (exclude identifier columns)
    conversion_errors = []
    for col in portal_numeric_cols:
        if col in df_processed.columns and col not in identifier_columns:
            original_values = df_processed[col].copy()
            df_processed[col] = df_processed[col].apply(lambda x: safe_convert_to_numeric(x, col))
            

            # Check for conversion issues
            conversion_count = (original_values != df_processed[col]).sum()
            # Convert Series to scalar if needed
            if hasattr(conversion_count, 'item'):
                conversion_count = conversion_count.item()
            elif hasattr(conversion_count, 'iloc'):
                conversion_count = conversion_count.iloc[0]
            
            if conversion_count > 0:
                conversion_errors.append(f"Column '{col}': {conversion_count} values converted from text to numeric")
                logger.warning(f"Column '{col}': {conversion_count} values were converted from text to numeric")
        elif col in identifier_columns:
            logger.info(f"Column '{col}' is an identifier column, keeping as text")
    
    # Check for missing required columns (exclude identifier columns that are used as index)
    missing_columns = []
    for col in required_columns:
        # Check if column exists in DataFrame columns or as index
        if col not in df_processed.columns and col not in df_processed.index.names:
            missing_columns.append(col)
    
    if missing_columns:
        error_msg = f"Missing required columns for {portal}: {missing_columns} | Available columns: {list(df_processed.columns)} | Index: {df_processed.index.names}"
        logger.error(error_msg)
        raise ValueError(error_msg)
    
    # Display conversion summary to user
    if conversion_errors:
        st.warning("⚠️ **Data Type Conversion Applied**")
        st.write("The following columns had text-formatted numbers that were automatically converted:")
        for error in conversion_errors:
            st.write(f"• {error}")
        st.info("💡 **Tip**: To avoid this in the future, ensure numeric columns in Excel are formatted as 'Number' instead of 'Text'.")
    
    return df_processed

def get_portal_columns(df, portal, required_cols, optional_cols=None):
    """Return validated portal columns, preserving optional columns when present."""
    optional_cols = optional_cols or []
    missing_cols = [col for col in required_cols if col not in df.columns]

    if missing_cols:
        available_cols = [str(col) for col in df.columns if col is not None]
        raise ValueError(
            f"Missing required columns for {portal}: {missing_cols}. "
            f"Available columns: {available_cols}"
        )

    return required_cols + [col for col in optional_cols if col in df.columns]

def process_excel_file(uploaded_file, target_profit, min_absolute_profit, portal, calculation_mode='discount', **kwargs):
    try:
        logger.info(f"Processing Excel file: {uploaded_file.name} for {portal} in {calculation_mode} mode")
        
        # Read the uploaded file
        wb = openpyxl.load_workbook(uploaded_file, data_only=False)
        ws = wb.active

        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        
        df_formulas = pd.DataFrame(data[1:], columns=data[0])
        df1 = df_formulas.iloc[:, :52]
        
        logger.info(f"Excel file loaded with {len(df1)} rows and {len(df1.columns)} columns")
        
        # Different column requirements based on portal
        if portal == 'Myntra':
            required_cols = ['ARTICLE NO', 'MRP', 'CP', 'GST',
                           'SHIPPING', 'COMMISSION %', 'FIXED FEE']
            optional_cols = ['REBATE TD', 'REBATE SP', 'REBATE VALUE']
            cols_to_use = get_portal_columns(df1, portal, required_cols, optional_cols)
            df2 = df1[cols_to_use]
            df3 = df2.copy()
            df3 = df3.set_index('ARTICLE NO')
        elif portal == 'Ajio':
            # For Ajio, only need basic columns for simple calculation
            # required_cols = ['ARTICLE NO', 'MRP', 'DISCOUNT %', 'stock status']
            required_cols = ['EAN', 'CP', 'Listing MRP', 'GST']
            df2 = df1[get_portal_columns(df1, portal, required_cols)]
            df3 = df2.copy()
            df3 = df3.set_index('EAN')
        elif portal == 'TataCliq':
            required_cols = ['SKU Code', 'CP', 'MRP', 'GST RATE']
            optional_cols = ['Processing fee']
            cols_to_use = get_portal_columns(df1, portal, required_cols, optional_cols)
            df2 = df1[cols_to_use]
            df3 = df2.copy()
            df3 = df3.set_index('SKU Code')
        elif portal == 'Nykaa':
            required_cols = ['SKU Code', 'MRP', 'cp', 'gst', 'shipping']
            df2 = df1[get_portal_columns(df1, portal, required_cols)]
            df3 = df2.copy()
            df3 = df3.set_index('SKU Code')
        elif portal == 'Pepperfry':
            required_cols = ['van', 'cp', 'mrp', 'gst']
            df2 = df1[get_portal_columns(df1, portal, required_cols)]
            df3 = df2.copy()
            df3 = df3.set_index('van')
        
        # Validate and convert data types to handle text-formatted numbers
        df3 = validate_and_convert_dataframe(df3, portal, required_cols)
        
        # Process the data based on calculation mode
        abs_profit_df = None
        if calculation_mode == 'mrp':
            logger.info(f"Starting MRP calculation for {portal} with {len(df3)} products")
            discount = kwargs.get('discount', 20)
            kwargs_without_discount = {k: v for k, v in kwargs.items() if k != 'discount'}
            result_df = build_mrp_table(df3, float(discount), float(target_profit), float(min_absolute_profit), portal, **kwargs_without_discount)
        elif calculation_mode == 'selling_price':
            logger.info(f"Starting selling price calculation for {portal} with {len(df3)} products")
            result_df = build_selling_price_table(df3, float(target_profit), float(min_absolute_profit), portal, **kwargs)
        else:
            logger.info(f"Starting profit calculation for {portal} with {len(df3)} products")
            result_df, abs_profit_df = build_profit_table(df3, float(target_profit), float(min_absolute_profit), portal, **kwargs)

        logger.info(f"Successfully completed processing for {portal}")
        return result_df, df_formulas, df3, abs_profit_df
        
    except Exception as e:
        logger.error(f"Error processing file: {str(e)} | File: {uploaded_file.name if uploaded_file else 'Unknown'} | Portal: {portal}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        st.error(f"Error processing file: {str(e)}")
        return None, None, None, None

def create_portal_page(portal_name, portal_emoji, calculation_info, data_format_info, additional_inputs=None, calculation_modes=None):
    """Create a page for a specific portal"""
    st.title(f"{portal_emoji} {portal_name} Pricing Analyzer")
    st.markdown(f"Upload your Excel file and set parameters to analyze pricing strategies for **{portal_name}**.")
    # Sidebar for inputs
    with st.sidebar:
        st.header("📥 Input Parameters")
        
        # Logging level control
        log_level = st.selectbox(
            "Console Log Level",
            ["INFO", "DEBUG", "WARNING", "ERROR"],
            index=0,
            help="Control the verbosity of console logs"
        )
        
        # Update logging level
        logging.getLogger().setLevel(getattr(logging, log_level))
        
        # Calculation mode selection (only for Myntra)
        calculation_mode = 'discount'  # Default mode for portals without an explicit mode selector
        if calculation_modes and portal_name == 'Myntra':
            default_mode_index = next(
                (idx for idx, mode in enumerate(calculation_modes) if "Selling Price" in mode),
                0
            )
            selected_mode = st.selectbox(
                "Calculation Mode",
                calculation_modes,
                index=default_mode_index,
                help="Choose between finding optimal discount or optimal MRP"
            )
            # Map user-friendly names to internal mode names
            if "MRP (Discount → MRP)" in selected_mode:
                calculation_mode = 'mrp'
            elif "Selling Price" in selected_mode:
                calculation_mode = 'selling_price'
            else:
                calculation_mode = 'discount'
        
        # Show detailed calculations toggle
        show_detailed_calc = st.checkbox(
            "Show Detailed Calculations",
            value=True,
            help="Display detailed calculation breakdown in expandable section below results table"
        )
        
        # File upload
        uploaded_file = st.file_uploader(
            "Choose Excel file",
            type=['xlsx', 'xls'],
            help=f"Upload your {portal_name} pricing data Excel file"
        )
        
        # Target profit input
        target_profit = st.number_input(
            "Target Profit Percentage (%)",
            min_value=0.0,
            max_value=100.0,
            value=15.0,
            step=0.1,
            help="Enter the minimum profit percentage you want to achieve"
        )
        
        # Minimum absolute profit input
        min_absolute_profit = st.number_input(
            "Minimum Absolute Profit (₹)",
            min_value=0.0,
            max_value=10000.0,
            value=100.0,
            step=10.0,
            help="Enter the minimum absolute profit amount in rupees"
        )
        
        # Additional inputs based on calculation mode
        extra_params = {'calculation_mode': calculation_mode}
        
        # For MRP calculation mode, add discount input
        if calculation_mode == 'mrp':
            extra_params['discount'] = st.number_input(
                "Fixed Discount Percentage (%)",
                min_value=0.0,
                max_value=99.0,
                value=20.0,
                step=1.0,
                help="Enter the fixed discount percentage to apply (MRP calculation mode)"
            )
        
        # Portal-specific additional inputs
        if additional_inputs:
            for input_config in additional_inputs:
                if input_config['type'] == 'number_input':
                    extra_params[input_config['key']] = st.number_input(
                        input_config['label'],
                        min_value=input_config.get('min_value', 0.0),
                        max_value=input_config.get('max_value', 100.0),
                        value=input_config.get('value', 0.0),
                        step=input_config.get('step', 0.1),
                        help=input_config.get('help', '')
                    )
                elif input_config['type'] == 'checkbox':
                    extra_params[input_config['key']] = st.checkbox(
                        input_config['label'],
                        value=input_config.get('default', False),
                        help=input_config.get('help', '')
                    )
                elif input_config['type'] == 'multiselect':
                    controlled_by = input_config.get('controlled_by')
                    if controlled_by and extra_params.get(controlled_by):
                        default_val = input_config.get('default_when_true', input_config.get('default', []))
                    else:
                        default_val = input_config.get('default', [])
                    selected = st.multiselect(
                        input_config['label'],
                        options=input_config.get('options', []),
                        default=default_val,
                        help=input_config.get('help', '')
                    )
                    extra_params[input_config['key']] = selected if selected else default_val

        # Remove internal control keys not needed downstream
        extra_params.pop('all_nine_endings', None)

        # Process button
        process_button = st.button("🚀 Process Data", type="primary")
    
    # Portal-specific information
    with st.expander(f"ℹ️ {portal_name} Calculation Details"):
        st.markdown(calculation_info)
    
    # Console log display area
    with st.expander("📋 Console Logs", expanded=False):
        st.info("Console logs will appear here when processing starts. Check the terminal/console where you ran the Streamlit app for detailed logs.")
        st.code("Logs are also saved to 'myntra_pricing.log' file in your project directory.", language="text")
    
    # Main content area
    state_key = f"{portal_name}_results"

    if uploaded_file is not None:
        st.success(f"✅ File uploaded: {uploaded_file.name}")

        if process_button:
            logger.info(f"User initiated processing for {portal_name} with target profit: {target_profit}%, min absolute profit: {min_absolute_profit}, mode: {calculation_mode}")
            with st.spinner(f"Processing your data for {portal_name}... This may take a few minutes."):
                result_df, original_df, processed_df, abs_profit_df = process_excel_file(uploaded_file, target_profit, min_absolute_profit, portal_name, **extra_params)

            if result_df is not None:
                st.session_state[state_key] = {
                    'result_df': result_df,
                    'original_df': original_df,
                    'processed_df': processed_df,
                    'abs_profit_df': abs_profit_df,
                    'calculation_mode': calculation_mode,
                    'extra_params': extra_params,
                    'target_profit': target_profit,
                    'min_absolute_profit': min_absolute_profit,
                }

        if state_key in st.session_state:
            saved = st.session_state[state_key]
            result_df = saved['result_df']
            original_df = saved['original_df']
            processed_df = saved['processed_df']
            abs_profit_df = saved['abs_profit_df']
            saved_mode = saved['calculation_mode']
            saved_extra_params = saved['extra_params']
            saved_target_profit = saved['target_profit']
            saved_min_absolute_profit = saved['min_absolute_profit']

            logger.info(f"Processing completed successfully for {portal_name}")
            st.success("✅ Processing completed!")

            # Display results
            if saved_mode == 'mrp':
                mode_text = "MRP Calculation"
            elif saved_mode == 'selling_price':
                mode_text = "Selling Price Analysis"
            else:
                mode_text = "9-Ending Selling Price Analysis"
            st.header(f"📈 Analysis Results - {portal_name} ({mode_text})")

            # Summary statistics
            col1, col2, col3, col4 = st.columns(4)

            with col1:
                total_products = len(result_df)
                st.metric("Total Products", total_products)

            if saved_mode == 'mrp':
                with col2:
                    products_with_solution = len(result_df[result_df['Status'] == 'Found'])
                    st.metric("Products with Solution", products_with_solution)

                with col3:
                    success_rate = (products_with_solution / total_products * 100) if total_products > 0 else 0
                    st.metric("Success Rate", f"{success_rate:.1f}%")

                with col4:
                    avg_mrp = result_df[result_df['Optimal MRP'].notna()]['Optimal MRP'].mean() if len(result_df[result_df['Optimal MRP'].notna()]) > 0 else 0
                    st.metric("Avg. Optimal MRP (₹)", f"₹{avg_mrp:.0f}")
            elif saved_mode == 'selling_price':
                with col2:
                    products_with_solution = len(result_df[result_df['Status'] == 'Found'])
                    st.metric("Products with Solution", products_with_solution)

                with col3:
                    success_rate = (products_with_solution / total_products * 100) if total_products > 0 else 0
                    st.metric("Success Rate", f"{success_rate:.1f}%")

                with col4:
                    avg_sp = result_df[result_df['Minimum Selling Price'].notna()]['Minimum Selling Price'].mean() if len(result_df[result_df['Minimum Selling Price'].notna()]) > 0 else 0
                    st.metric("Avg. Min SP (₹)", f"₹{avg_sp:.0f}")
            else:
                with col2:
                    products_with_target = len(result_df[result_df['Best Selling Price'].notna()])
                    st.metric("Products Meeting Target", products_with_target)

                with col3:
                    success_rate = (products_with_target / total_products * 100) if total_products > 0 else 0
                    st.metric("Success Rate", f"{success_rate:.1f}%")

                with col4:
                    avg_profit = result_df[result_df['Best Profit (₹)'] > 0]['Best Profit (₹)'].mean() if len(result_df[result_df['Best Profit (₹)'] > 0]) > 0 else 0
                    st.metric("Avg. Profit (₹)", f"₹{avg_profit:.0f}")

            # Display the results table
            st.subheader("📋 Detailed Results")
            st.dataframe(result_df, use_container_width=True)

            # Detailed calculations in expandable block
            if show_detailed_calc:
                st.markdown("---")
                if saved_mode == 'discount':
                    with st.expander("🔍 Detailed Calculations (First 2 Rows - Best 9-Ending Selling Price)", expanded=False):
                        display_detailed_calculations(processed_df, portal_name, result_df, **saved_extra_params)
                elif saved_mode == 'mrp':
                    with st.expander("🔍 Detailed Calculations (First 2 Rows - Optimal MRP)", expanded=False):
                        mrp_kwargs = saved_extra_params.copy()
                        mrp_kwargs['target_profit_percent'] = saved_target_profit
                        mrp_kwargs['min_absolute_profit'] = saved_min_absolute_profit
                        display_detailed_mrp_calculations(processed_df, portal_name, result_df, **mrp_kwargs)
                elif saved_mode == 'selling_price':
                    with st.expander("🔍 Detailed Calculations (First 2 Rows - Minimum Selling Price)", expanded=False):
                        display_detailed_selling_price_calculations(processed_df, portal_name, result_df)

                with st.expander("🔍 Detailed Calculations — Lookup by Article Number", expanded=False):
                    display_detailed_calculations_for_article(processed_df, portal_name, result_df, **saved_extra_params)

            # Download section
            st.header("💾 Download Results")

            # Create Excel file in memory
            current_time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                result_df.to_excel(writer, sheet_name=f'{portal_name} Analysis')
                if abs_profit_df is not None:
                    abs_profit_df.to_excel(writer, sheet_name='Profit (₹) per Discount')
                if original_df is not None:
                    original_df.to_excel(writer, sheet_name='Original Data', index=False)

            excel_data = output.getvalue()

            # Download button
            st.download_button(
                label="📥 Download Excel Report",
                data=excel_data,
                file_name=f'{portal_name.lower()}_pricing_analysis_{current_time}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            st.info(f"💡 The Excel file contains analysis for {portal_name} with your results and original data.")
    
    else:
        st.info("👆 Please upload an Excel file to get started.")

        # Template download
        template_columns = {
            'Myntra': ['ARTICLE NO', 'MRP', 'CP', 'GST',
                       'SHIPPING', 'COMMISSION %', 'FIXED FEE', 'REBATE TD', 'REBATE SP', 'REBATE VALUE'],
            'Ajio': ['EAN', 'CP', 'Listing MRP', 'GST'],
            'TataCliq': ['SKU Code', 'CP', 'MRP', 'GST RATE', 'Processing fee'],
            'Nykaa': ['SKU Code', 'MRP', 'cp', 'gst', 'shipping'],
            'Pepperfry': ['van', 'cp', 'mrp', 'gst'],
        }
        cols = template_columns.get(portal_name)
        if cols:
            template_output = io.BytesIO()
            pd.DataFrame(columns=cols).to_excel(template_output, index=False)
            st.download_button(
                label="📄 Download Template",
                data=template_output.getvalue(),
                file_name=f'{portal_name.lower()}_template.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        # Show sample data format
        with st.expander("📋 Expected Data Format"):
            st.markdown(data_format_info)

        # Show parameter explanation
        with st.expander("ℹ️ Parameter Explanation"):
            st.markdown(f"""
            **Target Profit Percentage (%)**: The minimum profit percentage you want to achieve on the selling price for {portal_name}.
            
            **Minimum Absolute Profit (₹)**: The minimum absolute profit amount in rupees that a product must generate to be considered viable.
            
            For example:
            - If Target Profit % = 15% and Min Absolute Profit = ₹100
            - A product must have both ≥15% profit AND ≥₹100 absolute profit to be considered as meeting the target
            """)

def myntra_page():
    """Myntra pricing analyzer page"""
    calculation_info = """
    **Myntra Calculation Includes:**
    - Customer shipping charges (formula-based)
    - Commission percentage (formula-based)
    - Fixed fees (formula-based)
    - Return fee (2% of selling price after logistics)
    - Marketing & packing cost (5% of selling price after logistics)
    - Complex cost structure with multiple variables
    
    **Calculation Modes:**
    - **Discount Analysis**: Find optimal discount percentage for given MRP to achieve target profit
    - **MRP Calculation**: Find optimal MRP for given discount percentage to achieve target profit
    - **Selling Price Analysis**: Find the minimum selling price using the selected price endings that achieves target profit
    """
    
    data_format_info = """
    **Required columns for Myntra:**
    - **ARTICLE NO**: Product article number
    - **MRP**: Maximum Retail Price (must be numeric)
    - **CP**: Cost price (must be numeric)
    - **GST**: GST percentage (must be numeric)
    - **SHIPPING**: Shipping charges formula
    - **COMMISSION %**: Commission percentage formula
    - **FIXED FEE**: Fixed fee formula

    **Optional columns:**
    - **REBATE TD**: Minimum discount % threshold to qualify for rebate
    - **REBATE SP**: Recommended selling price — rebate applies if selling price is below this
    - **REBATE VALUE**: Rebate percentage of selling price after logistics

    **Important Data Format Notes:**
    - Numeric columns (MRP, CP, GST) should be formatted as **Number** in Excel, not Text
    - Identifier columns (ARTICLE NO) should remain as **Text** format
    - If numeric columns are formatted as Text, the system will automatically convert them
    - Formula columns can contain Excel formulas like "IF(A1<500,50,0)"
    - Empty or invalid numeric values will be treated as 0
    """
    
    # Define calculation modes for Myntra
    calculation_modes = [
        "Discount (mrp → discount)",
        "MRP (discount → mrp)",
        "Selling Price (selected endings → minimum viable SP)"
    ]

    additional_inputs = [
        {
            'key': 'all_nine_endings',
            'type': 'checkbox',
            'label': 'All 9-Endings (Rakhi Mode)',
            'default': True,
            'help': 'On: selects all 9-endings by default. Off: uses 29, 49, 79, 99 only.'
        },
        {
            'key': 'price_endings',
            'type': 'multiselect',
            'label': 'Price Ending Options',
            'options': [9, 19, 29, 39, 49, 59, 69, 79, 89, 99],
            'default': [29, 49, 79, 99],
            'default_when_true': [9, 19, 29, 39, 49, 59, 69, 79, 89, 99],
            'controlled_by': 'all_nine_endings',
            'help': 'Round selling prices to end with one of these values (e.g. 249, 299)'
        }
    ]

    create_portal_page("Myntra", "🛍️", calculation_info, data_format_info, additional_inputs=additional_inputs, calculation_modes=calculation_modes)

def ajio_page():
    """Ajio pricing analyzer page"""
    calculation_info = """
    **Ajio Calculation:**
    - Selling price = Listing MRP - (Listing MRP × discount%)
    - GST value = Selling price × GST / 100
    - Commission = 35% of selling price
    - Commission GST = 18% of commission
    - Marketing = 10% of selling price
    - Total cost = Marketing + Total commission + GST value + CP
    - Profit = Selling price - Total cost
    - Profit percentage = Profit / Selling price × 100
    """
    
    data_format_info = """
    **Required columns for Ajio:**
    - **EAN**: Product EAN code
    - **CP**: Cost price (must be numeric)
    - **Listing MRP**: Maximum Retail Price (must be numeric)
    - **GST**: GST percentage (must be numeric)
    
    **Important Data Format Notes:**
    - Numeric columns (CP, Listing MRP, GST) should be formatted as **Number** in Excel, not Text
    - Identifier columns (EAN) should remain as **Text** format
    - If numeric columns are formatted as Text, the system will automatically convert them
    - Empty or invalid numeric values will be treated as 0
    - Ajio calculation uses fixed commission and marketing percentages in the current formula
    """
    
    # Additional inputs specific to Ajio
    additional_inputs = [
        {
            'type': 'number_input',
            'key': 'all_cost_percent',
            'label': 'All Cost Percentage (%)',
            'min_value': 0.0,
            'max_value': 100.0,
            'value': 42.0,
            'step': 0.1,
            'help': 'Enter the percentage of selling price that represents all costs (default: 42%)'
        }
    ]
    
    create_portal_page("Ajio", "🏪", calculation_info, data_format_info, additional_inputs)

def tatacliq_page():
    """TataCliq pricing analyzer page"""
    calculation_info = """
    **TataCliq Calculation Includes:**
    - Selling price = ROUND(MRP - (MRP × discount%), nearest 10) - 1
    - GST value = GST RATE × selling price / 100
    - Commission: ₹150 for orders < ₹500, 25% of selling price for orders ≥ ₹500
    - IGST = 18% of commission
    - Total fees = Commission + IGST
    - Shipping charge: ₹0 for orders < ₹500, ₹118 for orders ≥ ₹500
    - Marketing fees = 5% of selling price
    - Processing fee = Processing fee column value, or ₹49 if absent
    - Processing fee GST = 18% of processing fee
    - Total cost = GST value + shipping charge + total fees + CP + marketing fees + processing fee + processing fee GST
    - Profit = Selling price - Total cost
    """
    
    data_format_info = """
    **Required columns for TataCliq:**
    - **SKU Code**: Product SKU code
    - **CP**: Cost price (must be numeric)
    - **MRP**: Maximum Retail Price (must be numeric)
    - **GST RATE**: GST percentage rate (must be numeric)
    - **Processing fee**: Optional processing fee; defaults to ₹49 if missing
    
    **Important Data Format Notes:**
    - Numeric columns (CP, MRP, GST RATE, Processing fee) should be formatted as **Number** in Excel, not Text
    - Identifier columns (SKU Code) should remain as **Text** format
    - If numeric columns are formatted as Text, the system will automatically convert them
    - Empty or invalid numeric values will be treated as 0
    - TataCliq uses a complex calculation with variable commission and shipping charges based on order value
    """
    
    create_portal_page("TataCliq", "🛒", calculation_info, data_format_info)

def nykaa_page():
    """Nykaa pricing analyzer page"""
    calculation_info = """
    **Nykaa Calculation Includes:**
    - Selling price = MRP - (MRP × discount%)
    - GST value = GST rate × selling price / 100
    - Commission = 28% of selling price
    - Commission tax = 18% of commission
    - Total commission = Commission + Commission tax
    - Marketing fees = 10% of selling price
    - Payment fee = 0.8% of selling price
    - Payment fee GST = 18% of payment fee
    - Total cost = CP + GST value + shipping + marketing fees + total commission + payment fee + payment fee GST
    - Profit = Selling price - Total cost
    """
    
    data_format_info = """
    **Required columns for Nykaa:**
    - **SKU Code**: Product SKU code
    - **MRP**: Maximum Retail Price (must be numeric)
    - **cp**: Cost price (must be numeric)
    - **gst**: GST percentage rate (must be numeric)
    - **shipping**: Shipping charges (must be numeric)
    
    **Important Data Format Notes:**
    - Numeric columns (MRP, cp, gst, shipping) should be formatted as **Number** in Excel, not Text
    - Identifier columns (SKU Code) should remain as **Text** format
    - If numeric columns are formatted as Text, the system will automatically convert them
    - Empty or invalid numeric values will be treated as 0
    - Nykaa uses a fixed commission structure with 28% commission plus 18% tax on commission
    """
    
    create_portal_page("Nykaa", "💄", calculation_info, data_format_info)

def pepperfry_page():
    """Pepperfry pricing analyzer page"""
    calculation_info = """
    **Pepperfry Calculation:**
    - Selling price = MRP - (MRP × discount%)
    - GST value = GST × selling price / 100
    - Commission = 35% of selling price
    - Commission GST = 0% of commission
    - Total commission = Commission + Commission GST
    - Marketing = 10% of selling price
    - Total cost = Marketing + Total commission + GST value + CP
    - Profit = Selling price - Total cost
    - Profit percentage = Profit / Selling price × 100
    """
    
    data_format_info = """
    **Required columns for Pepperfry:**
    - **van**: Product identifier
    - **cp**: Cost price (must be numeric)
    - **mrp**: Maximum Retail Price (must be numeric)
    - **gst**: GST percentage (must be numeric)
    
    **Important Data Format Notes:**
    - Numeric columns (cp, mrp, gst) should be formatted as **Number** in Excel, not Text
    - Identifier columns (van) should remain as **Text** format
    - If numeric columns are formatted as Text, the system will automatically convert them
    - Empty or invalid numeric values will be treated as 0
    - Pepperfry calculation uses 35% commission and 10% marketing
    """
    
    # Additional inputs specific to Pepperfry (same as Ajio)
    additional_inputs = [
        {
            'type': 'number_input',
            'key': 'all_cost_percent',
            'label': 'All Cost Percentage (%)',
            'min_value': 0.0,
            'max_value': 100.0,
            'value': 42.0,
            'step': 0.1,
            'help': 'Enter the percentage of selling price that represents all costs (default: 42%)'
        }
    ]
    
    create_portal_page("Pepperfry", "🪑", calculation_info, data_format_info, additional_inputs)

def commission_analysis_page():
    """Commission comparison page for Aapno Rajasthan, Tayhaa, Aura, MYEH"""
    st.title("📊 Commission Comparison")
    st.markdown("Upload a commission Excel file to compare rates across **Aapno Rajasthan, Tayhaa, Aura** and **MYEH**.")

    uploaded = st.file_uploader("Upload Commission Excel", type=["xlsx", "xls"])
    if uploaded is None:
        st.info("Upload the commission file to get started.")
        return

    try:
        df_raw = pd.read_excel(uploaded, sheet_name="Sheet1", header=None)
    except Exception as e:
        st.error(f"Could not read file: {e}")
        return

    price_bands = ["0-300", "300-500", "500-1000", "1000-2000", "2000-100000"]
    brand_cols = {
        "Aapno":  [11, 12, 13, 14, 15],
        "Tayhaa": [26, 27, 28, 29, 30],
        "Aura":   [39, 40, 41, 42, 43],
        "MYEH":   [52, 53, 54, 55, 56],
    }
    brands = list(brand_cols.keys())

    # Parse data
    rows = []
    for _, row in df_raw.iloc[2:].iterrows():
        article = row[1]
        if pd.isna(article) or str(article).strip() in ["", "-", "NaN"]:
            continue
        entry = {"Article Type": str(article).strip()}
        for brand, cols in brand_cols.items():
            for band, col in zip(price_bands, cols):
                val = row[col]
                try:
                    entry[f"{brand}_{band}"] = float(val) if not pd.isna(val) else None
                except Exception:
                    entry[f"{brand}_{band}"] = None
        rows.append(entry)

    import numpy as np
    data = pd.DataFrame(rows).drop_duplicates("Article Type").reset_index(drop=True)

    # Build summary
    summary_rows = []
    for _, row in data.iterrows():
        cat = row["Article Type"]
        for band in price_bands:
            vals = {b: row[f"{b}_{band}"] for b in brands}
            valid = {b: v for b, v in vals.items() if v is not None and not np.isnan(v)}
            if not valid:
                continue
            min_val = min(valid.values())
            winners = [b for b, v in valid.items() if v == min_val]
            summary_rows.append({
                "Category": cat,
                "Price Band": band,
                "Cheapest Brand": " / ".join(winners),
                "Lowest Commission (%)": min_val,
                **{b: vals[b] for b in brands},
            })

    summary = pd.DataFrame(summary_rows)

    # ── Sidebar filters ───────────────────────────────────────────────────
    with st.sidebar:
        st.header("🔍 Filters")
        selected_bands = st.multiselect("Price Bands", price_bands, default=price_bands)
        all_cats = sorted(summary["Category"].unique())
        selected_cats = st.multiselect("Categories", all_cats, default=all_cats)
        selected_brands = st.multiselect("Show brands", brands, default=brands)

    filtered = summary[
        summary["Price Band"].isin(selected_bands) &
        summary["Category"].isin(selected_cats)
    ]

    # ── Win count cards ───────────────────────────────────────────────────
    st.subheader("Which brand is cheapest most often?")
    counts = {b: 0 for b in brands}
    for _, row in filtered.iterrows():
        for b in row["Cheapest Brand"].split(" / "):
            if b.strip() in counts:
                counts[b.strip()] += 1

    cols = st.columns(4)
    brand_colors = {"Aapno": "🟢", "Tayhaa": "🔵", "Aura": "🟠", "MYEH": "🟣"}
    for col, (brand, cnt) in zip(cols, sorted(counts.items(), key=lambda x: -x[1])):
        col.metric(f"{brand_colors.get(brand, '')} {brand}", f"{cnt} times cheapest")

    st.markdown("---")

    # ── Main table ────────────────────────────────────────────────────────
    st.subheader("Commission Rates by Category & Price Band")

    display_cols = ["Category", "Price Band", "Cheapest Brand", "Lowest Commission (%)"] + [b for b in brands if b in selected_brands]
    display_df = filtered[display_cols].copy()

    def highlight_cheapest(row):
        styles = [""] * len(row)
        cheapest = [b.strip() for b in str(row.get("Cheapest Brand", "")).split(" / ")]
        for i, col in enumerate(row.index):
            if col in brands:
                val = row[col]
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    styles[i] = "color: #aaa"
                elif col in cheapest:
                    styles[i] = "background-color: #C6EFCE; font-weight: bold"
                else:
                    styles[i] = "background-color: #FFC7CE"
        return styles

    st.dataframe(
        display_df.style.apply(highlight_cheapest, axis=1).format(
            {b: lambda x: f"{x:.0f}%" if x is not None and not np.isnan(x) else "N/A" for b in brands},
            na_rep="N/A"
        ),
        use_container_width=True,
        height=500,
    )

    # ── Pivot: cheapest brand per category across bands ───────────────────
    st.markdown("---")
    st.subheader("Cheapest Brand at a Glance")
    pivot = filtered.pivot_table(
        index="Category", columns="Price Band", values="Cheapest Brand", aggfunc="first"
    ).reindex(columns=price_bands)
    st.dataframe(pivot, use_container_width=True)

    # ── Download ──────────────────────────────────────────────────────────
    st.markdown("---")
    import io
    buf = io.BytesIO()
    full_pivot = summary.pivot_table(
        index="Category", columns="Price Band", values="Cheapest Brand", aggfunc="first"
    ).reindex(columns=price_bands).reset_index()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Lowest Per Band", index=False)
        full_pivot.to_excel(writer, sheet_name="Cheapest Brand Pivot", index=False)
        pd.DataFrame([{"Brand": b, "Times Cheapest": counts[b]} for b in brands]).to_excel(writer, sheet_name="Brand Win Count", index=False)
        data.to_excel(writer, sheet_name="Raw Rates", index=False)

    # Apply colours to "Lowest Per Band" sheet
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    green  = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red    = PatternFill("solid", fgColor="FFC7CE")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)

    wb = load_workbook(buf)
    ws = wb["Lowest Per Band"]

    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Find brand column indices (1-based) from header row
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    brand_col_indices = {b: header.index(b) + 1 for b in brands if b in header}
    cheapest_col = header.index("Cheapest Brand") + 1 if "Cheapest Brand" in header else None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cheapest_brands = []
        if cheapest_col:
            val = ws.cell(row[0].row, cheapest_col).value
            cheapest_brands = [b.strip() for b in str(val).split(" / ")] if val else []
        for brand, col_i in brand_col_indices.items():
            cell = ws.cell(row[0].row, col_i)
            if cell.value is None or str(cell.value).strip() == "":
                continue
            if brand in cheapest_brands:
                cell.fill = green if len(cheapest_brands) == 1 else yellow
            else:
                cell.fill = red

    # Auto-width all sheets
    for sheet_name in wb.sheetnames:
        ws_cur = wb[sheet_name]
        for col in ws_cur.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws_cur.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

    buf2 = io.BytesIO()
    wb.save(buf2)
    st.download_button("⬇️ Download Full Analysis", buf2.getvalue(), "commission_analysis.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def main():
    logger.info("Starting Multi-Portal Pricing Analyzer application")

    st.set_page_config(
        page_title="Multi-Portal Pricing Analyzer",
        page_icon="📊",
        layout="wide"
    )

    # Define pages
    myntra_page_obj = st.Page(myntra_page, title="Myntra Portal", icon="🛍️")
    ajio_page_obj = st.Page(ajio_page, title="Ajio Portal", icon="🏪")
    tatacliq_page_obj = st.Page(tatacliq_page, title="TataCliq Portal", icon="🛒")
    nykaa_page_obj = st.Page(nykaa_page, title="Nykaa Portal", icon="💄")
    pepperfry_page_obj = st.Page(pepperfry_page, title="Pepperfry Portal", icon="🪑")
    commission_page_obj = st.Page(commission_analysis_page, title="Commission Comparison", icon="📊")

    # Create navigation
    pg = st.navigation([myntra_page_obj, ajio_page_obj, tatacliq_page_obj, nykaa_page_obj, pepperfry_page_obj, commission_page_obj])

    # Run the selected page
    pg.run()

if __name__ == "__main__":
    main() 
